from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Funcionario, ApontamentoFuncionario, FechamentoSemanal
from .models import ApontamentoDiarioLote, FuncionarioLote, RegistroProducao, FotoApontamento
from .models import HistoricoAlteracaoEtapa
from .forms import (
    FuncionarioForm, ApontamentoForm, FechamentoForm,
    ApontamentoDiarioCabecalhoForm, ApontamentoDiarioLoteForm,
)
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from urllib.parse import urlencode
from apps.obras.models import (
    Obra, Etapa,
    Etapa1Fundacao, Etapa2Estrutura, Etapa3Instalacoes,
    Etapa4Acabamentos, Etapa5Finalizacao, EtapaHistorico,
)
from django.db import IntegrityError
from django.db.models import Sum, Count, Q, Avg, F, Value
from django.db.models.functions import Replace
import datetime
from decimal import Decimal, InvalidOperation
from collections import defaultdict
import calendar
from apps.obras.templatetags.obras_extras import brl
from django.db import transaction
from django.views.decorators.http import require_GET, require_http_methods
import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


# ==================== ETAPA ITEMS HELPERS ====================

# Metadata for fields in each etapa detail model
ETAPA_FIELDS_META = {
    1: {
        'related_name': 'fundacao',
        'model_class': Etapa1Fundacao,
        'fields': [
            ('limpeza_terreno', 'boolean', 'Limpeza do Terreno'),
            ('instalacao_energia_agua', 'boolean', 'Instalação de Energia e Água'),
            ('marcacao_escavacao_inicio', 'date', 'Marcação e Escavação (início)'),
            ('marcacao_escavacao_conclusao', 'date', 'Marcação e Escavação (conclusão)'),
            ('locacao_ferragem_inicio', 'date', 'Locação de Ferragem (início)'),
            ('locacao_ferragem_conclusao', 'date', 'Locação de Ferragem (conclusão)'),
            ('aterro_contrapiso_inicio', 'date', 'Aterro e Contrapiso (início)'),
            ('aterro_contrapiso_conclusao', 'date', 'Aterro e Contrapiso (conclusão)'),
            ('fiadas_respaldo_inicio', 'date', '8 Fiadas até Respaldo (início)'),
            ('fiadas_respaldo_conclusao', 'date', '8 Fiadas até Respaldo (conclusão)'),
            ('levantar_alicerce_percentual', 'decimal', 'Levantar Alicerce, Reboco e Impermeabilizar (%)'),
        ]
    },
    2: {
        'related_name': 'estrutura',
        'model_class': Etapa2Estrutura,
        'fields': [
            ('montagem_laje_inicio', 'date', 'Montagem da Laje (início)'),
            ('montagem_laje_conclusao', 'date', 'Montagem da Laje (conclusão)'),
            ('cobertura_inicio', 'date', 'Cobertura Completa (início)'),
            ('cobertura_conclusao', 'date', 'Cobertura Completa (conclusão)'),
            ('platibanda_blocos', 'integer', 'Platibanda (Unidades de Blocos)'),
        ]
    },
    3: {
        'related_name': 'instalacoes',
        'model_class': Etapa3Instalacoes,
        'fields': [
            ('reboco_externo_m2', 'decimal', 'Reboco Externo (m²)'),
            ('reboco_interno_m2', 'decimal', 'Reboco Interno (m²)'),
            ('instalacao_portais', 'boolean', 'Instalação de Portais'),
            ('agua_fria', 'boolean', 'Água Fria'),
            ('esgoto', 'boolean', 'Esgoto'),
            ('fluvial', 'boolean', 'Fluvial'),
        ]
    },
    4: {
        'related_name': 'acabamentos',
        'model_class': Etapa4Acabamentos,
        'fields': [
            ('portas_janelas', 'boolean', 'Portas e Janelas'),
            ('pintura_externa_1demao_inicio', 'date', 'Pintura Externa 1ª Demão (início)'),
            ('pintura_externa_1demao_conclusao', 'date', 'Pintura Externa 1ª Demão (conclusão)'),
            ('pintura_interna_1demao_inicio', 'date', 'Pintura Interna 1ª Demão (início)'),
            ('pintura_interna_1demao_conclusao', 'date', 'Pintura Interna 1ª Demão (conclusão)'),
            ('assentamento_piso_inicio', 'date', 'Assentamento de Piso (início)'),
            ('assentamento_piso_conclusao', 'date', 'Assentamento de Piso (conclusão)'),
        ]
    },
    5: {
        'related_name': 'finalizacao',
        'model_class': Etapa5Finalizacao,
        'fields': [
            ('pintura_externa_2demao_inicio', 'date', 'Pintura Externa 2ª Demão (início)'),
            ('pintura_externa_2demao_conclusao', 'date', 'Pintura Externa 2ª Demão (conclusão)'),
            ('pintura_interna_2demao_inicio', 'date', 'Pintura Interna 2ª Demão (início)'),
            ('pintura_interna_2demao_conclusao', 'date', 'Pintura Interna 2ª Demão (conclusão)'),
            ('loucas_metais', 'boolean', 'Louças e Metais'),
            ('eletrica', 'boolean', 'Elétrica'),
        ]
    },
}


def _get_etapa_detail_obj(etapa, create=True):
    """Get (or optionally create) the detail model instance for an etapa."""
    meta = ETAPA_FIELDS_META.get(etapa.numero_etapa)
    if not meta:
        return None
    model_class = meta['model_class']
    related = meta['related_name']
    try:
        return getattr(etapa, related)
    except model_class.DoesNotExist:
        if create:
            return model_class.objects.create(etapa=etapa)
        return None


def _get_etapa_items(etapa):
    """Return a list of dicts with field metadata + current values for an etapa."""
    meta = ETAPA_FIELDS_META.get(etapa.numero_etapa)
    if not meta:
        return []
    detail_obj = _get_etapa_detail_obj(etapa, create=False)
    items = []
    for field_name, field_type, label in meta['fields']:
        current_value = None
        if detail_obj:
            current_value = getattr(detail_obj, field_name, None)
        if current_value is None:
            current_value = False if field_type == 'boolean' else 0
        # Normalize date to ISO string for the UI
        if field_type == 'date' and current_value:
            try:
                current_value = current_value.isoformat()
            except Exception:
                current_value = ''
        items.append({
            'name': field_name,
            'type': field_type,
            'label': label,
            'value': str(current_value) if isinstance(current_value, Decimal) else current_value,
        })
    return items


def _update_etapa_items_from_post(etapa, post_data):
    """Update etapa detail model fields from POST data (ADDITIVE logic) and recalculate obra progress.

    Returns a list of change description strings (empty list if nothing changed).

    - boolean fields: OR logic (once marked True, stays True)
    - integer/decimal fields: posted value is ADDED to the current value
    """
    meta = ETAPA_FIELDS_META.get(etapa.numero_etapa)
    if not meta:
        return []
    detail_obj = _get_etapa_detail_obj(etapa, create=True)
    if not detail_obj:
        return []

    changed = False
    alteracoes = []  # list of "Label: antes → depois" strings
    for field_name, field_type, label in meta['fields']:
        key = f'item_{field_name}'
        if field_type == 'boolean':
            # Hidden "0" + checkbox "1" pattern — OR logic
            val = post_data.get(key, '0')
            new_val = val in ('1', 'on', 'true', 'True')
            current_val = getattr(detail_obj, field_name, False)
            final_val = current_val or new_val  # once True, stays True
            if current_val != final_val:
                setattr(detail_obj, field_name, final_val)
                changed = True
                alteracoes.append(f"{label}: {'Sim' if current_val else 'Não'} → {'Sim' if final_val else 'Não'}")
        elif field_type == 'integer':
            raw = post_data.get(key, '')
            if raw != '':
                try:
                    increment = int(raw)
                    if increment != 0:
                        current_val = getattr(detail_obj, field_name, 0) or 0
                        novo_val = current_val + increment
                        setattr(detail_obj, field_name, novo_val)
                        changed = True
                        alteracoes.append(f"{label}: {current_val} → {novo_val} (+{increment})")
                except (ValueError, TypeError):
                    pass
        elif field_type == 'decimal':
            raw = post_data.get(key, '')
            if raw != '':
                try:
                    increment = Decimal(raw)
                    if increment != Decimal('0'):
                        current_val = getattr(detail_obj, field_name, Decimal('0')) or Decimal('0')
                        novo_val = current_val + increment
                        setattr(detail_obj, field_name, novo_val)
                        changed = True
                        alteracoes.append(f"{label}: {current_val} → {novo_val} (+{increment})")
                except (ValueError, TypeError, InvalidOperation):
                    pass
        elif field_type == 'date':
            raw = post_data.get(key, '')
            if raw:
                try:
                    new_date = datetime.date.fromisoformat(raw)
                    current_val = getattr(detail_obj, field_name, None)
                    if current_val != new_date:
                        setattr(detail_obj, field_name, new_date)
                        changed = True
                        antes = current_val.strftime('%d/%m/%Y') if current_val else '—'
                        alteracoes.append(f"{label}: {antes} → {new_date.strftime('%d/%m/%Y')}")
                except ValueError:
                    pass

    if changed:
        detail_obj.save()
        # Recalculate obra overall progress
        etapa.obra.calcular_percentual()

    return alteracoes


def _registrar_historico_apontamento(etapa, apontamento, request, is_update=False, etapa_items_changes=None):
    """Registra no histórico da etapa todas as informações do apontamento."""
    ap = apontamento
    obra = ap.obra
    func = ap.funcionario

    acao = 'Atualizado' if is_update else 'Criado'

    linhas = [
        f"Funcionário: {func.nome_completo} ({func.get_funcao_display()})",
        f"Obra: {obra.nome}",
        f"Endereço: {obra.endereco}",
    ]
    if obra.cliente:
        linhas.append(f"Cliente: {obra.cliente.nome}")
    linhas.append(f"Data: {ap.data.strftime('%d/%m/%Y')}")
    linhas.append(f"Horas Trabalhadas: {ap.horas_trabalhadas}h")
    linhas.append(f"Clima: {ap.get_clima_display()}")
    if ap.metragem_executada and ap.metragem_executada > 0:
        linhas.append(f"Metragem Executada: {ap.metragem_executada} m²")
    linhas.append(f"Valor Diária: R$ {ap.valor_diaria}")
    if ap.houve_ociosidade:
        linhas.append(f"⚠️ Ociosidade: {ap.observacao_ociosidade or 'Sem justificativa'}")
    if ap.houve_retrabalho:
        linhas.append(f"⚠️ Retrabalho: {ap.motivo_retrabalho or 'Sem motivo informado'}")
    if ap.observacoes:
        linhas.append(f"Observações: {ap.observacoes}")

    # Append etapa items changes if any
    if etapa_items_changes:
        linhas.append("")
        linhas.append("── Itens da Etapa ──")
        linhas.extend(etapa_items_changes)

    usuario = request.user if request.user and request.user.is_authenticated else None

    EtapaHistorico.objects.create(
        etapa=etapa,
        usuario=usuario,
        origem=f'Apontamento {acao}',
        descricao='\n'.join(linhas)
    )


def _calcular_producao_lote_por_campos(valores_producao_dia):
    """
    Calcula producao total do dia e infere unidade com base nos campos da etapa.
    """
    if not valores_producao_dia:
        return Decimal('0.00'), 'blocos'

    totais = {
        'blocos': Decimal('0.00'),
        'm2': Decimal('0.00'),
        'percentual': Decimal('0.00'),
        'unidades': Decimal('0.00'),
    }

    for campo, valor in valores_producao_dia.items():
        try:
            v = Decimal(str(valor))
        except Exception:
            continue
        if v <= 0:
            continue

        campo_lower = (campo or '').lower()
        if 'm2' in campo_lower:
            totais['m2'] += v
        elif 'percentual' in campo_lower:
            totais['percentual'] += v
        elif 'bloco' in campo_lower or 'fiadas' in campo_lower:
            totais['blocos'] += v
        else:
            totais['unidades'] += v

    ativos = [(u, t) for u, t in totais.items() if t > 0]
    if not ativos:
        return Decimal('0.00'), 'blocos'

    # Se houve mistura de unidades, nao soma entre si no campo escalar do lote
    if len(ativos) > 1:
        return Decimal('0.00'), 'blocos'

    unidade, total = ativos[0]
    return total, unidade


def _obter_detalhes_etapa(etapa):
    """Busca/cria o model de detalhes da etapa."""
    if not etapa:
        return None
    numero_etapa = etapa.numero_etapa
    if numero_etapa == 1:
        detalhes, _ = Etapa1Fundacao.objects.get_or_create(etapa=etapa)
    elif numero_etapa == 2:
        detalhes, _ = Etapa2Estrutura.objects.get_or_create(etapa=etapa)
    elif numero_etapa == 3:
        detalhes, _ = Etapa3Instalacoes.objects.get_or_create(etapa=etapa)
    elif numero_etapa == 4:
        detalhes, _ = Etapa4Acabamentos.objects.get_or_create(etapa=etapa)
    elif numero_etapa == 5:
        detalhes, _ = Etapa5Finalizacao.objects.get_or_create(etapa=etapa)
    else:
        detalhes = None
    return detalhes


def _processar_campos_etapa_payload(etapa, campos_payload, usuario=None):
    """
    Aplica payload de campos da etapa (formato rascunho do frontend).
    Campos numéricos são incrementais; boolean/date sobrescrevem.
    """
    detalhes = _obter_detalhes_etapa(etapa)
    if not detalhes:
        return {}, []

    campos_atualizados = []
    valores_producao_dia = {}

    for chave, valor in (campos_payload or {}).items():
        campo_nome = str(chave or '').strip()
        if not campo_nome:
            continue
        if campo_nome.startswith('campo_'):
            campo_nome = campo_nome.replace('campo_', '', 1)

        if not hasattr(detalhes, campo_nome):
            continue

        try:
            field = detalhes._meta.get_field(campo_nome)
            valor_anterior = getattr(detalhes, campo_nome, None)
            field_type = field.get_internal_type()

            if field_type == 'BooleanField':
                if isinstance(valor, bool):
                    novo_valor = valor
                else:
                    novo_valor = str(valor).strip().lower() in ('1', 'true', 'on', 'sim')
                if valor_anterior != novo_valor:
                    setattr(detalhes, campo_nome, novo_valor)
                    campos_atualizados.append(f"{field.verbose_name}: {'✓' if novo_valor else '✗'}")
                continue

            valor_str = '' if valor is None else str(valor).strip()
            if not valor_str or valor_str in ('0', '0.00'):
                continue

            if field_type == 'DecimalField':
                novo_valor = Decimal(valor_str)
                if novo_valor <= 0:
                    continue

                valor_anterior_decimal = valor_anterior if valor_anterior else Decimal('0.00')
                valor_final = valor_anterior_decimal + novo_valor

                tem_max_100 = any(
                    hasattr(v, 'limit_value') and v.limit_value == Decimal('100.00')
                    for v in field.validators
                )
                if tem_max_100 and valor_final > Decimal('100.00'):
                    diferenca = Decimal('100.00') - valor_anterior_decimal
                    if diferenca <= 0:
                        continue
                    valor_final = Decimal('100.00')
                    valores_producao_dia[campo_nome] = diferenca
                    campos_atualizados.append(f"{field.verbose_name}: +{diferenca} (total: 100.00)")
                else:
                    valores_producao_dia[campo_nome] = novo_valor
                    campos_atualizados.append(f"{field.verbose_name}: +{novo_valor} (total: {valor_final})")

                setattr(detalhes, campo_nome, valor_final)
                continue

            if field_type in ['IntegerField', 'PositiveIntegerField']:
                novo_valor = int(Decimal(valor_str))
                if novo_valor <= 0:
                    continue

                valor_anterior_int = valor_anterior if valor_anterior else 0
                valor_final = valor_anterior_int + novo_valor
                setattr(detalhes, campo_nome, valor_final)
                valores_producao_dia[campo_nome] = novo_valor
                campos_atualizados.append(f"{field.verbose_name}: +{novo_valor} (total: {valor_final})")
                continue

            if field_type == 'DateField':
                novo_valor = datetime.datetime.strptime(valor_str, '%Y-%m-%d').date()
                if valor_anterior != novo_valor:
                    setattr(detalhes, campo_nome, novo_valor)
                    campos_atualizados.append(f"{field.verbose_name}: {novo_valor.strftime('%d/%m/%Y')}")

        except Exception:
            continue

    if campos_atualizados:
        detalhes.save()
        try:
            EtapaHistorico.objects.create(
                etapa=etapa,
                origem='Apontamento em Lote',
                descricao="📝 Campos atualizados via apontamento em lote:\n" + "\n".join([f"  • {c}" for c in campos_atualizados]),
                usuario=usuario
            )
        except Exception:
            pass

    return valores_producao_dia, campos_atualizados


def _criar_lote_por_payload(base_data, etapa, request, funcionarios_ids, horas_trabalhadas_list, campos_payload, fotos=None):
    lote = ApontamentoDiarioLote.objects.create(
        obra=base_data['obra'],
        data=base_data['data'],
        etapa=etapa,
        clima=base_data['clima'],
        houve_ociosidade=base_data['houve_ociosidade'],
        observacao_ociosidade=base_data['observacao_ociosidade'],
        houve_retrabalho=base_data['houve_retrabalho'],
        motivo_retrabalho=base_data['motivo_retrabalho'],
        possui_placa=base_data['possui_placa'],
        observacoes=base_data['observacoes'],
        criado_por=request.user,
    )

    funcionarios_criados = 0
    for i, func_id in enumerate(funcionarios_ids):
        try:
            funcionario = Funcionario.objects.get(pk=func_id, ativo=True)
            horas = Decimal(horas_trabalhadas_list[i]) if i < len(horas_trabalhadas_list) else Decimal('8.0')
            if funcionario.funcao == 'fiscal':
                horas = Decimal('0.0')
            elif horas <= Decimal('0.0'):
                horas = Decimal('8.0')
            FuncionarioLote.objects.create(
                lote=lote,
                funcionario=funcionario,
                horas_trabalhadas=horas
            )
            funcionarios_criados += 1
        except (Funcionario.DoesNotExist, ValueError, InvalidOperation):
            continue

    if funcionarios_criados == 0:
        lote.delete()
        return None, 0, 0

    valores_producao_dia, _campos_atualizados = _processar_campos_etapa_payload(
        etapa=etapa,
        campos_payload=campos_payload,
        usuario=request.user
    )

    producao_total_dia, unidade_dia = _calcular_producao_lote_por_campos(valores_producao_dia)
    lote.producao_total = producao_total_dia
    lote.unidade_medida = unidade_dia
    lote.save(update_fields=['producao_total', 'unidade_medida'])

    lote._valores_dia = valores_producao_dia
    apontamentos_criados = lote.gerar_apontamentos_individuais() or 0

    for foto in (fotos or []):
        FotoApontamento.objects.create(
            apontamento_lote=lote,
            obra=lote.obra,
            etapa=lote.etapa,
            data_foto=lote.data,
            foto=foto
        )

    return lote, funcionarios_criados, apontamentos_criados


@login_required
def funcionario_list(request):
    """Lista funcionários"""
    funcionarios = Funcionario.objects.filter(ativo=True).order_by('nome_completo')

    # Filtro por função
    funcoes_choices = list(Funcionario.FUNCAO_CHOICES)
    funcoes_validas = {key for key, _ in funcoes_choices}
    funcao_filter = request.GET.get('funcao', '')
    if funcao_filter in funcoes_validas:
        funcionarios = funcionarios.filter(funcao=funcao_filter)

    # Busca por nome ou CPF (ignora pontos/traços na comparação)
    busca = request.GET.get('q', '').strip()
    if busca:
        import re
        busca_lower = busca.lower()
        funcoes_match = [
            key for key, label in funcoes_choices
            if busca_lower in key.lower() or busca_lower in str(label).lower()
        ]
        digits = re.sub(r'\D', '', busca)
        if digits:
            # remove '.' '-' and spaces from cpf field for comparison
            funcionarios = funcionarios.annotate(
                cpf_digits=Replace(
                    Replace(
                        Replace(F('cpf'), Value('.'), Value('')),
                        Value('-'), Value('')
                    ),
                    Value(' '), Value('')
                )
            ).filter(
                Q(nome_completo__icontains=busca)
                | Q(cpf_digits__icontains=digits)
                | Q(funcao__in=funcoes_match)
            )
        else:
            funcionarios = funcionarios.filter(
                Q(nome_completo__icontains=busca)
                | Q(funcao__in=funcoes_match)
            )

    # Contadores
    total_ativos = Funcionario.objects.filter(ativo=True).count()
    total_pedreiros = Funcionario.objects.filter(ativo=True, funcao='pedreiro').count()
    total_serventes = Funcionario.objects.filter(ativo=True, funcao='servente').count()
    total_resultado = funcionarios.count()
    funcoes_ativas_counts = dict(
        Funcionario.objects.filter(ativo=True)
        .values_list('funcao')
        .annotate(total=Count('id'))
    )
    funcoes_filtro = [
        {
            'key': key,
            'label': label,
            'count': funcoes_ativas_counts.get(key, 0),
        }
        for key, label in funcoes_choices
    ]

    # Paginação
    per_page = request.GET.get('per_page', '15')
    if per_page not in ('10', '15', '20'):
        per_page = '15'
    per_page = int(per_page)
    paginator = Paginator(funcionarios, per_page)
    page = request.GET.get('page')
    try:
        funcionarios_page = paginator.page(page)
    except PageNotAnInteger:
        funcionarios_page = paginator.page(1)
    except EmptyPage:
        funcionarios_page = paginator.page(paginator.num_pages)

    context = {
        'funcionarios': funcionarios_page,
        'page_obj': funcionarios_page,
        'title': 'Funcionários',
        'busca': busca,
        'funcao_filter': funcao_filter,
        'total_ativos': total_ativos,
        'total_pedreiros': total_pedreiros,
        'total_serventes': total_serventes,
        'total_resultado': total_resultado,
        'per_page': per_page,
        'funcoes_filtro': funcoes_filtro,
    }
    return render(request, 'funcionarios/funcionario_list.html', context)


@login_required
def funcionario_detail(request, pk):
    """Hub completo do funcionário — mostra tudo num só lugar com filtro de período."""
    funcionario = get_object_or_404(Funcionario, pk=pk)

    # Filtro de período com presets
    hoje = datetime.date.today()
    preset = request.GET.get('preset', '30dias')
    data_inicio_str = request.GET.get('data_inicio', '')
    data_fim_str = request.GET.get('data_fim', '')

    if data_inicio_str and data_fim_str:
        try:
            data_inicio = datetime.date.fromisoformat(data_inicio_str)
            data_fim = datetime.date.fromisoformat(data_fim_str)
            preset = 'custom'
        except ValueError:
            data_inicio = hoje - datetime.timedelta(days=30)
            data_fim = hoje
    else:
        if preset == 'semana':
            # Início da semana (segunda)
            data_inicio = hoje - datetime.timedelta(days=hoje.weekday())
            data_fim = hoje
        elif preset == 'mes':
            data_inicio = hoje.replace(day=1)
            data_fim = hoje
        elif preset == 'quinzena':
            data_inicio = hoje - datetime.timedelta(days=15)
            data_fim = hoje
        elif preset == '90dias':
            data_inicio = hoje - datetime.timedelta(days=90)
            data_fim = hoje
        else:  # 30dias
            data_inicio = hoje - datetime.timedelta(days=30)
            data_fim = hoje

    # --- Apontamentos no período ---
    apontamentos = ApontamentoFuncionario.objects.filter(
        funcionario=funcionario,
        data__gte=data_inicio,
        data__lte=data_fim
    ).select_related('obra', 'etapa').order_by('-data')

    # KPIs
    kpis = apontamentos.aggregate(
        total_dias=Count('data', distinct=True),
        total_horas=Sum('horas_trabalhadas'),
        total_valor=Sum('valor_diaria'),
        total_metragem=Sum('metragem_executada'),
        dias_ociosidade=Count('data', filter=Q(houve_ociosidade=True), distinct=True),
        dias_retrabalho=Count('data', filter=Q(houve_retrabalho=True), distinct=True),
    )
    total_dias = kpis['total_dias'] or 0
    total_horas = kpis['total_horas'] or Decimal('0.0')
    total_valor = kpis['total_valor'] or Decimal('0.00')
    total_metragem = kpis['total_metragem'] or Decimal('0.00')
    dias_ociosidade = kpis['dias_ociosidade'] or 0
    dias_retrabalho = kpis['dias_retrabalho'] or 0
    taxa_ociosidade = round(dias_ociosidade / total_dias * 100, 1) if total_dias else 0
    taxa_retrabalho = round(dias_retrabalho / total_dias * 100, 1) if total_dias else 0
    media_horas = round(total_horas / total_dias, 1) if total_dias else Decimal('0.0')

    # --- Obras trabalhadas no período ---
    obras_periodo = (
        apontamentos
        .values('obra__pk', 'obra__nome')
        .annotate(
            dias=Count('data', distinct=True),
            horas=Sum('horas_trabalhadas'),
            valor=Sum('valor_diaria'),
            metragem=Sum('metragem_executada'),
        )
        .order_by('-dias')
    )

    # --- Etapas trabalhadas ---
    etapas_periodo = (
        apontamentos
        .filter(etapa__isnull=False)
        .values('etapa__numero_etapa')
        .annotate(
            dias=Count('data', distinct=True),
            metragem=Sum('metragem_executada'),
        )
        .order_by('etapa__numero_etapa')
    )
    ETAPA_NOMES = {
        1: 'Fundação', 2: 'Estrutura', 3: 'Instalações',
        4: 'Acabamentos', 5: 'Finalização',
    }
    for ep in etapas_periodo:
        ep['etapa_nome'] = ETAPA_NOMES.get(ep['etapa__numero_etapa'], f"Etapa {ep['etapa__numero_etapa']}")

    # --- Fechamentos no período ---
    fechamentos = FechamentoSemanal.objects.filter(
        funcionario=funcionario,
        data_inicio__lte=data_fim,
        data_fim__gte=data_inicio,
    ).order_by('-data_inicio')

    # --- Últimos apontamentos (últimos 15) ---
    ultimos_apontamentos = apontamentos[:15]

    context = {
        'funcionario': funcionario,
        'title': funcionario.nome_completo,
        # Período
        'preset': preset,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        # KPIs
        'total_dias': total_dias,
        'total_horas': total_horas,
        'total_valor': total_valor,
        'total_metragem': total_metragem,
        'dias_ociosidade': dias_ociosidade,
        'dias_retrabalho': dias_retrabalho,
        'taxa_ociosidade': taxa_ociosidade,
        'taxa_retrabalho': taxa_retrabalho,
        'media_horas': media_horas,
        # Dados
        'obras_periodo': obras_periodo,
        'etapas_periodo': etapas_periodo,
        'fechamentos': fechamentos,
        'ultimos_apontamentos': ultimos_apontamentos,
    }
    return render(request, 'funcionarios/funcionario_detail.html', context)


@login_required
def funcionario_create(request):
    """Cadastra novo funcionário"""
    if request.method == 'POST':
        form = FuncionarioForm(request.POST, request.FILES)
        if form.is_valid():
            funcionario = form.save()
            messages.success(request, 'Funcionário criado com sucesso.')
            return redirect('funcionarios:funcionario_detail', pk=funcionario.pk)
        else:
            messages.error(request, 'Corrija os erros no formulário.')
    else:
        form = FuncionarioForm()
    return render(request, 'funcionarios/funcionario_form.html', {'form': form, 'title': 'Novo Funcionário'})


@login_required
def funcionario_update(request, pk):
    """Atualiza funcionário"""
    funcionario = get_object_or_404(Funcionario, pk=pk)
    if request.method == 'POST':
        form = FuncionarioForm(request.POST, request.FILES, instance=funcionario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Funcionário atualizado com sucesso.')
            return redirect('funcionarios:funcionario_detail', pk=funcionario.pk)
        else:
            messages.error(request, 'Corrija os erros no formulário.')
    else:
        form = FuncionarioForm(instance=funcionario)
    return render(request, 'funcionarios/funcionario_form.html', {'form': form, 'funcionario': funcionario, 'title': 'Editar Funcionário'})


@login_required
def funcionario_inativar(request, pk):
    """Inativa um funcionário"""
    funcionario = get_object_or_404(Funcionario, pk=pk)
    return render(request, 'funcionarios/funcionario_inativar.html', {
        'funcionario': funcionario,
        'title': 'Inativar Funcionário'
    })


# ==================== APONTAMENTOS ====================

def _aplicar_filtros_apontamentos(request, queryset=None):
    qs = (queryset or ApontamentoFuncionario.objects.all()).select_related(
        'funcionario', 'obra', 'etapa'
    ).order_by('-data', '-created_at')

    data = request.GET.get('data')
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')
    funcionario_id = request.GET.get('funcionario')
    obra_id = request.GET.get('obra')
    etapa_id = request.GET.get('etapa')
    clima = request.GET.get('clima')
    ocorrencia = request.GET.get('ocorrencia')

    if data_inicio_str or data_fim_str:
        try:
            data_inicio = datetime.date.fromisoformat(data_inicio_str) if data_inicio_str else None
            data_fim = datetime.date.fromisoformat(data_fim_str) if data_fim_str else None
        except ValueError:
            data_inicio = data_fim = None

        if data_inicio and data_fim:
            qs = qs.filter(data__gte=data_inicio, data__lte=data_fim)
        elif data_inicio:
            qs = qs.filter(data__gte=data_inicio)
        elif data_fim:
            qs = qs.filter(data__lte=data_fim)
    elif data:
        qs = qs.filter(data=data)

    if funcionario_id:
        qs = qs.filter(funcionario_id=funcionario_id)
    if obra_id:
        qs = qs.filter(obra_id=obra_id)
    if etapa_id:
        qs = qs.filter(etapa_id=etapa_id)
    if clima:
        qs = qs.filter(clima=clima)
    if ocorrencia == 'ociosidade':
        qs = qs.filter(houve_ociosidade=True)
    elif ocorrencia == 'retrabalho':
        qs = qs.filter(houve_retrabalho=True)

    return qs


def _texto_status_apontamento(apontamento):
    if apontamento.houve_retrabalho:
        return 'Retrabalho'
    if apontamento.houve_ociosidade:
        return 'Ociosidade'
    return 'OK'


def _resumo_exportacao_apontamentos(request, apontamentos):
    total_valor = apontamentos.aggregate(total_valor=Sum('valor_diaria'))['total_valor'] or Decimal('0')
    data = request.GET.get('data')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if data_inicio and data_fim:
        periodo = f'{datetime.date.fromisoformat(data_inicio):%d/%m/%Y} a {datetime.date.fromisoformat(data_fim):%d/%m/%Y}'
    elif data_inicio:
        periodo = f'A partir de {datetime.date.fromisoformat(data_inicio):%d/%m/%Y}'
    elif data_fim:
        periodo = f'Ate {datetime.date.fromisoformat(data_fim):%d/%m/%Y}'
    elif data:
        periodo = f'{datetime.date.fromisoformat(data):%d/%m/%Y}'
    else:
        periodo = 'Todos'

    return {
        'periodo': periodo,
        'valor_total': total_valor,
    }


def _exportar_apontamentos_excel(request, apontamentos):
    resumo = _resumo_exportacao_apontamentos(request, apontamentos)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Apontamentos'
    moeda_fmt = 'R$ #,##0.00'
    header_fill = PatternFill(fill_type='solid', fgColor='0D6EFD')
    total_fill = PatternFill(fill_type='solid', fgColor='E2F0D9')

    ws.append(['Relatorio de Diarias'])
    ws.append([f"Periodo: {resumo['periodo']}"])
    ws.append(['Valor Total', float(resumo['valor_total'])])
    ws.append([f"Gerado em: {timezone.localtime():%d/%m/%Y %H:%M}"])
    ws.append([])
    ws.append([
        'Funcionario', 'Funcao', 'Obra', 'Etapa', 'Data',
        'Horas', 'Clima', 'Valor', 'Status',
    ])

    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'].font = Font(bold=True)
    ws['A3'].font = Font(bold=True)
    ws['B3'].number_format = moeda_fmt
    ws['A1'].alignment = Alignment(horizontal='left')
    ws['A2'].alignment = Alignment(horizontal='left')
    ws['A3'].alignment = Alignment(horizontal='left')
    ws['B3'].alignment = Alignment(horizontal='left')

    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True, color='FFFFFF')

    for ap in apontamentos:
        ws.append([
            ap.funcionario.nome_completo,
            ap.funcionario.get_funcao_display(),
            ap.obra.nome if ap.obra else '-',
            str(ap.etapa) if ap.etapa else '-',
            ap.data.strftime('%d/%m/%Y'),
            float(ap.horas_trabalhadas or 0),
            ap.get_clima_display(),
            float(ap.valor_diaria or 0),
            _texto_status_apontamento(ap),
        ])

    ws.append([])
    ws.append(['Valor Total', '', '', '', '', '', '', float(resumo['valor_total']), ''])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = total_fill

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row - 2):
        row[5].alignment = Alignment(horizontal='center')
        row[7].number_format = moeda_fmt
        row[7].alignment = Alignment(horizontal='right')
        row[8].alignment = Alignment(horizontal='center')

    ws.cell(row=ws.max_row, column=8).number_format = moeda_fmt
    ws.freeze_panes = 'A7'

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="apontamentos_{timezone.now():%Y%m%d_%H%M%S}.xlsx"'
    )
    wb.save(response)
    return response


def _exportar_apontamentos_pdf(request, apontamentos):
    resumo = _resumo_exportacao_apontamentos(request, apontamentos)
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )
    styles = getSampleStyleSheet()
    cell_style = styles['BodyText']
    cell_style.fontName = 'Helvetica'
    cell_style.fontSize = 8
    cell_style.leading = 9

    linhas = [[
        'Funcionario', 'Obra', 'Etapa', 'Data', 'Horas', 'Clima', 'Valor', 'Status',
    ]]
    for ap in apontamentos:
        linhas.append([
            Paragraph(ap.funcionario.nome_completo, cell_style),
            Paragraph(ap.obra.nome if ap.obra else '-', cell_style),
            Paragraph(str(ap.etapa) if ap.etapa else '-', cell_style),
            Paragraph(ap.data.strftime('%d/%m/%Y'), cell_style),
            Paragraph(f'{ap.horas_trabalhadas}h', cell_style),
            Paragraph(ap.get_clima_display(), cell_style),
            Paragraph(brl(ap.valor_diaria or 0), cell_style),
            Paragraph(_texto_status_apontamento(ap), cell_style),
        ])

    tabela = Table(
        linhas,
        repeatRows=1,
        colWidths=[3.5 * cm, 3.4 * cm, 6.0 * cm, 2.0 * cm, 1.8 * cm, 2.0 * cm, 2.6 * cm, 2.2 * cm],
    )
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    story = [
        Paragraph('Relatorio de Diarias', styles['Title']),
        Paragraph(f"Periodo: {resumo['periodo']}", styles['Normal']),
        Paragraph(f"Valor Total: {brl(resumo['valor_total'])}", styles['Normal']),
        Paragraph(f'Gerado em {timezone.localtime():%d/%m/%Y %H:%M}', styles['Normal']),
        Spacer(1, 12),
        tabela,
    ]
    doc.build(story)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="apontamentos_{timezone.now():%Y%m%d_%H%M%S}.pdf"'
    )
    response.write(buffer.getvalue())
    buffer.close()
    return response


@login_required
def apontamento_list(request):
    """Lista apontamentos com filtros avancados"""
    qs = _aplicar_filtros_apontamentos(request)

    exportar = request.GET.get('export')
    if exportar == 'excel':
        return _exportar_apontamentos_excel(request, qs)
    if exportar == 'pdf':
        return _exportar_apontamentos_pdf(request, qs)

    totais = qs.aggregate(
        total_horas=Sum('horas_trabalhadas'),
        total_valor=Sum('valor_diaria'),
        total_registros=Count('id'),
        total_diarias=Count('data', distinct=True),
        total_ociosidade=Count('data', filter=Q(houve_ociosidade=True), distinct=True),
        total_retrabalho=Count('data', filter=Q(houve_retrabalho=True), distinct=True),
    )

    paginator = Paginator(qs, 20)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    params = request.GET.copy()
    if 'page' in params:
        params.pop('page')
    if 'export' in params:
        params.pop('export')
    querystring = params.urlencode()

    context = {
        'apontamentos': page_obj.object_list,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'paginator': paginator,
        'querystring': querystring,
        'totais': totais,
        'funcionarios': Funcionario.objects.filter(ativo=True).order_by('nome_completo'),
        'obras': Obra.objects.filter(
            ativo=True,
            status__in=['planejamento', 'em_andamento']
        ).order_by('nome'),
        'title': 'Apontamentos'
    }
    return render(request, 'funcionarios/apontamento_list.html', context)


@login_required
def apontamento_create(request, funcionario_id=None):
    """Cria apontamento individual"""
    if request.method == 'POST':
        form = ApontamentoForm(request.POST, request.FILES, funcionario_id=funcionario_id)
        if form.is_valid():
            ap = form.save(commit=False)
            ap.valor_diaria = ap.funcionario.valor_diaria
            # ✅ SEMPRE criar novo registro (permite múltiplos apontamentos mesmo func/dia/obra)
            ap.save()
            is_update = False

            # ---- Auto-update obra: save etapa items from POST ----
            etapa_items_changes = []
            if ap.etapa and request.POST.get('items_etapa_id'):
                etapa_items_changes = _update_etapa_items_from_post(ap.etapa, request.POST)
                if etapa_items_changes:
                    messages.info(request, f'📊 Progresso da etapa "{ap.etapa}" atualizado automaticamente.')

            # ---- Registrar histórico no EtapaHistorico ----
            if ap.etapa:
                _registrar_historico_apontamento(ap.etapa, ap, request, is_update=is_update, etapa_items_changes=etapa_items_changes)

            # ========== PROCESSAR FOTOS ==========
            fotos_uploaded = request.FILES.getlist('fotos')
            for foto in fotos_uploaded:
                FotoApontamento.objects.create(
                    apontamento_individual=ap,
                    obra=ap.obra,
                    etapa=ap.etapa,
                    data_foto=ap.data,
                    foto=foto
                )
            
            if fotos_uploaded:
                messages.info(request, f'📷 {len(fotos_uploaded)} foto(s) anexada(s)!')
            # ========================================

            # Notificação de retrabalho/ociosidade
            if ap.houve_retrabalho:
                messages.warning(request, f'⚠️ RETRABALHO registrado para {ap.funcionario.nome_completo}.')
            if ap.houve_ociosidade:
                messages.warning(request, f'⚠️ OCIOSIDADE registrada para {ap.funcionario.nome_completo}.')
            messages.success(request, 'Apontamento salvo com sucesso.')
            return redirect('funcionarios:apontamento_list')
        else:
            messages.error(request, 'Corrija os erros no formulário.')
    else:
        form = ApontamentoForm(funcionario_id=funcionario_id)
    return render(request, 'funcionarios/apontamento_form.html', {'form': form, 'title': 'Novo Apontamento'})


# DESCONTINUADO em 21/02/2026: usar apontamento_lote_create no lugar.
# Esta view foi removida em favor do sistema de apontamento em lote
# que permite registrar múltiplos funcionários de uma vez com indicadores de produção.
# @login_required
# def apontamento_diario(request):
#     """
#     Registro diário em lote — simplificado:
#     1. Fiscal seleciona obra + data + clima
#     2. Tabela mostra todos os funcionários ativos → marca quem trabalhou
#     3. Salva todos de uma vez com um clique
#     """
#     cab_form = ApontamentoDiarioCabecalhoForm(request.GET or None)

# REMOVIDO em 21/02/2026: usar apontamento_lote_create no lugar
# @login_required
# def apontamento_diario(request):
#     """
#     DESCONTINUADO: Usar apontamento_lote_create no lugar.
#     Mantida temporariamente para evitar erros de importação.
#     Redireciona para o novo sistema de apontamento em lote.
#     """
#     from django.shortcuts import redirect as _redirect
#     return _redirect('funcionarios:apontamento_lote_create')


def _apontamento_diario_legado(request):
    """Código legado da view apontamento_diario — não utilizado."""
    cab_form = ApontamentoDiarioCabecalhoForm(request.GET or None)
    obra = None
    data = None
    clima = None
    apontamentos_existentes = []
    funcionarios_disponiveis = []
    etapas = []

    # Step 1: Check if obra + data are set
    obra_id = request.GET.get('obra') or request.POST.get('obra')
    data_str = request.GET.get('data') or request.POST.get('data_apontamento')
    clima_sel = request.GET.get('clima') or request.POST.get('clima_apontamento')

    if obra_id and data_str:
        try:
            obra = Obra.objects.get(
                pk=obra_id,
                ativo=True,
                status__in=['planejamento', 'em_andamento']
            )
            data = datetime.date.fromisoformat(data_str)
            clima = clima_sel or 'sol'
        except (Obra.DoesNotExist, ValueError):
            pass

    if obra and data:
        etapas = Etapa.objects.filter(
            obra=obra,
            status='em_andamento',
        ).order_by('numero_etapa')
        apontamentos_existentes = ApontamentoFuncionario.objects.filter(
            obra=obra, data=data
        ).select_related('funcionario', 'etapa')

        # IDs já apontados nesta obra/data
        apontados_ids = set(apontamentos_existentes.values_list('funcionario_id', flat=True))

        # Lista de funcionários ativos disponíveis para apontar
        todos_funcionarios = Funcionario.objects.filter(ativo=True).order_by('nome_completo')
        for f in todos_funcionarios:
            funcionarios_disponiveis.append({
                'obj': f,
                'ja_apontado': f.pk in apontados_ids,
                'bloqueado': False,
            })

        # Set initial values for cab_form
        cab_form = ApontamentoDiarioCabecalhoForm(initial={
            'obra': obra.pk, 'data': data, 'clima': clima
        })

        if request.method == 'POST' and 'salvar_lote' in request.POST:
            # Processar apontamento em lote
            func_ids = request.POST.getlist('func_ids')
            criados = 0
            atualizados = 0
            erros = 0

            for func_id in func_ids:
                try:
                    func = Funcionario.objects.get(pk=func_id, ativo=True)
                except Funcionario.DoesNotExist:
                    erros += 1
                    continue

                # Pega dados da linha
                horas = request.POST.get(f'horas_{func_id}', '8.0')
                etapa_id = request.POST.get(f'etapa_{func_id}', '')

                try:
                    horas_dec = Decimal(horas)
                    if horas_dec < Decimal('0.5'):
                        horas_dec = Decimal('8.0')
                except (InvalidOperation, ValueError):
                    horas_dec = Decimal('8.0')
                if func.funcao == 'fiscal':
                    horas_dec = Decimal('0.0')

                etapa_obj = None
                if etapa_id:
                    try:
                        etapa_obj = Etapa.objects.get(
                            pk=etapa_id,
                            obra=obra,
                            status='em_andamento',
                        )
                    except Etapa.DoesNotExist:
                        pass

                # ✅ SEMPRE criar novo (permite múltiplos turnos/períodos no mesmo dia)
                ap_novo = ApontamentoFuncionario.objects.create(
                    funcionario=func,
                    obra=obra,
                    etapa=etapa_obj,
                    data=data,
                    horas_trabalhadas=horas_dec,
                    clima=clima,
                    valor_diaria=func.valor_diaria,
                )
                criados += 1
                # Registrar histórico
                if etapa_obj:
                    _registrar_historico_apontamento(etapa_obj, ap_novo, request, is_update=False)

            if criados or atualizados:
                msg_parts = []
                if criados:
                    msg_parts.append(f'{criados} apontamento(s) criado(s)')
                if atualizados:
                    msg_parts.append(f'{atualizados} atualizado(s)')
                messages.success(request, ', '.join(msg_parts) + '.')
            if erros:
                messages.warning(request, f'{erros} funcionário(s) não puderam ser apontados.')

            return redirect(f'{request.path}?obra={obra.pk}&data={data.isoformat()}&clima={clima}')

    context = {
        'cab_form': cab_form,
        'obra': obra,
        'data': data,
        'clima': clima,
        'etapas': etapas,
        'apontamentos_existentes': apontamentos_existentes,
        'funcionarios_disponiveis': funcionarios_disponiveis,
        'title': 'Registro Diário'
    }
    return render(request, 'funcionarios/apontamento_diario.html', context)


@login_required
def apontamento_delete(request, pk):
    """Remove um apontamento"""
    ap = get_object_or_404(ApontamentoFuncionario, pk=pk)
    obra_id = ap.obra_id
    data = ap.data
    funcionario_id = ap.funcionario_id
    etapa_id = ap.etapa_id

    def _limpar_registros_producao_orfaos(func_id, ob_id, dt, et_id):
        """
        Remove RegistroProducao órfão após exclusão de apontamento individual.
        Só exclui quando não existe mais apontamento para o mesmo funcionário/obra/data/etapa.
        """
        qs_ap = ApontamentoFuncionario.objects.filter(
            funcionario_id=func_id,
            obra_id=ob_id,
            data=dt,
        )
        qs_reg = RegistroProducao.objects.filter(
            funcionario_id=func_id,
            obra_id=ob_id,
            data=dt,
        )

        if et_id is None:
            qs_ap = qs_ap.filter(etapa__isnull=True)
            qs_reg = qs_reg.filter(etapa__isnull=True)
        else:
            qs_ap = qs_ap.filter(etapa_id=et_id)
            qs_reg = qs_reg.filter(etapa_id=et_id)

        if not qs_ap.exists():
            qs_reg.delete()

    if request.method == 'POST':
        # Registrar exclusão no histórico da etapa antes de deletar
        if ap.etapa:
            func = ap.funcionario
            obra = ap.obra
            linhas = [
                f"Funcionário: {func.nome_completo} ({func.get_funcao_display()})",
                f"Obra: {obra.nome}",
                f"Endereço: {obra.endereco}",
            ]
            if obra.cliente:
                linhas.append(f"Cliente: {obra.cliente.nome}")
            linhas.append(f"Data: {ap.data.strftime('%d/%m/%Y')}")
            linhas.append(f"Horas Trabalhadas: {ap.horas_trabalhadas}h")
            linhas.append(f"Clima: {ap.get_clima_display()}")
            if ap.metragem_executada and ap.metragem_executada > 0:
                linhas.append(f"Metragem Executada: {ap.metragem_executada} m²")
            linhas.append(f"Valor Diária: R$ {ap.valor_diaria}")

            usuario = request.user if request.user and request.user.is_authenticated else None
            EtapaHistorico.objects.create(
                etapa=ap.etapa,
                usuario=usuario,
                origem='Apontamento Excluído',
                descricao='\n'.join(linhas)
            )

        ap.delete()
        _limpar_registros_producao_orfaos(funcionario_id, obra_id, data, etapa_id)
        messages.success(request, 'Apontamento removido.')
        # Redirect back to diario if referer suggests it
        next_url = request.POST.get('next', '')
        if next_url:
            return redirect(next_url)
    return redirect('funcionarios:apontamento_list')


# ==================== FECHAMENTOS ====================

@login_required
def fechamento_list(request):
    """Lista fechamentos agrupados por semana"""
    from django.db.models import Sum, Count, Q, Min, Max

    # Agrupar por semana (data_inicio, data_fim)
    semanas_qs = (
        FechamentoSemanal.objects
        .values('data_inicio', 'data_fim')
        .annotate(
            total_funcionarios=Count('id'),
            total_dias=Sum('total_dias'),
            total_valor=Sum('total_valor'),
            total_ociosidade=Sum('dias_ociosidade'),
            total_retrabalho=Sum('dias_retrabalho'),
            qtd_fechados=Count('id', filter=Q(status='fechado')),
            qtd_pagos=Count('id', filter=Q(status='pago')),
        )
        .order_by('-data_inicio')
    )

    semanas = list(semanas_qs)

    # Calcular status geral de cada semana
    for s in semanas:
        if s['qtd_pagos'] == s['total_funcionarios']:
            s['status_geral'] = 'pago'
        elif s['qtd_pagos'] > 0:
            s['status_geral'] = 'parcial'
        else:
            s['status_geral'] = 'fechado'

    context = {
        'semanas': semanas,
        'title': 'Fechamentos Semanais',
    }
    return render(request, 'funcionarios/fechamento_list.html', context)


@login_required
def set_theme(request):
    """Endpoint to persist user's theme preference (light/dark)."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)
    theme = request.POST.get('theme')
    variant = request.POST.get('variant')
    if theme not in ('light', 'dark'):
        return JsonResponse({'status': 'error', 'message': 'Invalid theme'}, status=400)
    profile = getattr(request.user, 'profile', None)
    if not profile:
        from .models import UserProfile
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
    profile.theme_preference = theme
    if variant in dict(getattr(profile, 'THEME_VARIANT_CHOICES', [])) or variant in ('default','soft','gray','blue'):
        profile.theme_variant = variant
    profile.save()
    return JsonResponse({'status': 'ok', 'theme': theme})


@login_required
def fechamento_semana_detail(request, data_inicio):
    """Detalhe de uma semana: lista todos os funcionários e seus fechamentos."""
    try:
        dt_inicio = datetime.date.fromisoformat(data_inicio)
    except ValueError:
        messages.error(request, 'Data inválida.')
        return redirect('funcionarios:fechamento_list')

    dt_fim = dt_inicio + datetime.timedelta(days=5)  # seg a sáb

    fechamentos = (
        FechamentoSemanal.objects
        .filter(data_inicio=dt_inicio)
        .select_related('funcionario')
        .order_by('funcionario__nome_completo')
    )

    if not fechamentos.exists():
        messages.warning(request, 'Nenhum fechamento encontrado para esta semana.')
        return redirect('funcionarios:fechamento_list')

    dt_fim_real = fechamentos.first().data_fim

    # Totais da semana
    from django.db.models import Sum, Count, Q
    totais = fechamentos.aggregate(
        total_funcionarios=Count('id'),
        total_dias=Sum('total_dias'),
        total_valor=Sum('total_valor'),
        total_ociosidade=Sum('dias_ociosidade'),
        total_retrabalho=Sum('dias_retrabalho'),
        qtd_fechados=Count('id', filter=Q(status='fechado')),
        qtd_pagos=Count('id', filter=Q(status='pago')),
    )

    # Filtro por status
    status_filter = request.GET.get('status')
    if status_filter in ('fechado', 'pago'):
        fechamentos = fechamentos.filter(status=status_filter)

    # Busca por nome
    busca = request.GET.get('q', '').strip()
    if busca:
        fechamentos = fechamentos.filter(funcionario__nome_completo__icontains=busca)

    # Filtro por dia específico ou intervalo de datas (filtra via apontamentos)
    dia_filter = request.GET.get('dia', '').strip()
    dia_inicio_filter = request.GET.get('dia_inicio', '').strip()
    dia_fim_filter = request.GET.get('dia_fim', '').strip()

    if dia_filter:
        # Dia específico: só mostra funcionários que trabalharam nesse dia
        try:
            dt_dia = datetime.date.fromisoformat(dia_filter)
            func_ids = ApontamentoFuncionario.objects.filter(
                data=dt_dia,
                data__gte=dt_inicio,
                data__lte=dt_fim_real,
            ).values_list('funcionario_id', flat=True)
            fechamentos = fechamentos.filter(funcionario_id__in=func_ids)
        except ValueError:
            pass
    elif dia_inicio_filter or dia_fim_filter:
        # Intervalo de datas dentro da semana
        try:
            dt_di = datetime.date.fromisoformat(dia_inicio_filter) if dia_inicio_filter else dt_inicio
            dt_df = datetime.date.fromisoformat(dia_fim_filter) if dia_fim_filter else dt_fim_real
            # Limitar ao intervalo da semana
            dt_di = max(dt_di, dt_inicio)
            dt_df = min(dt_df, dt_fim_real)
            func_ids = ApontamentoFuncionario.objects.filter(
                data__gte=dt_di,
                data__lte=dt_df,
            ).values_list('funcionario_id', flat=True)
            fechamentos = fechamentos.filter(funcionario_id__in=func_ids)
        except ValueError:
            pass

    # Gerar lista de dias da semana para os filtros rápidos
    dias_semana = []
    DIAS_NOMES = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
    d = dt_inicio
    while d <= dt_fim_real:
        dias_semana.append({
            'data': d,
            'nome': DIAS_NOMES[d.weekday()],
            'iso': d.isoformat(),
        })
        d += datetime.timedelta(days=1)

    # Buscar apontamentos da semana para mostrar dia/obra por funcionário
    apontamentos_semana = (
        ApontamentoFuncionario.objects
        .filter(data__gte=dt_inicio, data__lte=dt_fim_real)
        .select_related('obra')
        .order_by('data')
    )
    # Agrupar por funcionário: lista de {data, obra_nome}
    apts_por_func = defaultdict(list)
    for apt in apontamentos_semana:
        apts_por_func[apt.funcionario_id].append({
            'data': apt.data,
            'obra': apt.obra.nome if apt.obra else '—',
        })

    # Converter fechamentos para lista para poder anotar
    fechamentos_list = list(fechamentos)
    for f in fechamentos_list:
        f.apontamentos_semana = apts_por_func.get(f.funcionario_id, [])

    context = {
        'fechamentos': fechamentos_list,
        'data_inicio': dt_inicio,
        'data_fim': dt_fim_real,
        'totais': totais,
        'status_filter': status_filter or '',
        'busca': busca,
        'dia_filter': dia_filter,
        'dia_inicio_filter': dia_inicio_filter,
        'dia_fim_filter': dia_fim_filter,
        'dias_semana': dias_semana,
        'title': f'Semana {dt_inicio.strftime("%d/%m/%Y")} a {dt_fim_real.strftime("%d/%m/%Y")}',
    }
    return render(request, 'funcionarios/fechamento_semana_detail.html', context)


@login_required
def fechamento_semana_pagar(request, data_inicio):
    """Marca todos os fechamentos da semana como pagos (POST).

    Apenas usuários com permissão de alteração ou staff podem executar.
    """
    if request.method != 'POST':
        return redirect('funcionarios:fechamento_semana_detail', data_inicio=data_inicio)

    if not (request.user.is_staff or request.user.has_perm('funcionarios.change_fechamentosemanal')):
        return HttpResponseForbidden('Permissão negada')

    try:
        dt_inicio = datetime.date.fromisoformat(data_inicio)
    except ValueError:
        messages.error(request, 'Data inválida.')
        return redirect('funcionarios:fechamento_list')

    fechamentos_qs = FechamentoSemanal.objects.filter(data_inicio=dt_inicio)
    if not fechamentos_qs.exists():
        messages.warning(request, 'Nenhum fechamento encontrado para esta semana.')
        return redirect('funcionarios:fechamento_semana_detail', data_inicio=data_inicio)

    hoje = timezone.now().date()
    with transaction.atomic():
        updated = fechamentos_qs.exclude(status='pago').update(status='pago', data_pagamento=hoje)

    messages.success(request, f'{updated} fechamento(s) marcado(s) como pago.')
    return redirect('funcionarios:fechamento_semana_detail', data_inicio=data_inicio)


@login_required
def fechamento_create(request):
    """Cria um fechamento com período flexível e calcula os totais"""
    if request.method == 'POST':
        form = FechamentoForm(request.POST)
        if form.is_valid():
            f = form.save(commit=False)
            f.status = 'fechado'
            # Check duplicate
            existing = FechamentoSemanal.objects.filter(
                funcionario=f.funcionario,
                data_inicio=f.data_inicio,
                data_fim=f.data_fim
            ).first()
            if existing:
                messages.warning(request, 'Já existe um fechamento para esse funcionário e período.')
                return redirect('funcionarios:fechamento_detail', pk=existing.pk)
            try:
                f.save()
            except IntegrityError:
                messages.error(request, 'Erro ao salvar: já existe um fechamento para esse funcionário e período.')
                return redirect('funcionarios:fechamento_list')
            f.calcular_totais()
            messages.success(request, 'Fechamento criado e totais calculados.')
            return redirect('funcionarios:fechamento_detail', pk=f.pk)
        else:
            messages.error(request, 'Corrija os erros no formulário.')
    else:
        form = FechamentoForm()
    return render(request, 'funcionarios/fechamento_form.html', {'form': form, 'title': 'Novo Fechamento'})


@login_required
def fechamento_detail(request, pk):
    """Detalhes de um fechamento semanal"""
    fechamento = get_object_or_404(FechamentoSemanal, pk=pk)
    apontamentos = fechamento.get_apontamentos()
    obras_etapas = fechamento.get_obras_etapas()
    context = {
        'fechamento': fechamento,
        'apontamentos': apontamentos,
        'obras_etapas': obras_etapas,
        'title': f'Fechamento - {fechamento.funcionario.nome_completo}'
    }
    return render(request, 'funcionarios/fechamento_detail.html', context)


@login_required
def fechamento_delete(request, pk):
    """Deleta um fechamento semanal com confirmação"""
    fechamento = get_object_or_404(FechamentoSemanal, pk=pk)
    
    if request.method == 'POST':
        # Validar se o fechamento já foi pago (proteção)
        if fechamento.status == 'pago':
            messages.error(request, 'Não é possível excluir um fechamento que já foi pago.')
            return redirect('funcionarios:fechamento_detail', pk=pk)
        
        data_inicio = fechamento.data_inicio
        nome_funcionario = fechamento.funcionario.nome_completo
        
        try:
            fechamento.delete()
            messages.success(
                request,
                f'Fechamento de {nome_funcionario} ({data_inicio.strftime("%d/%m/%Y")}) foi excluído com sucesso.'
            )
            return redirect('funcionarios:fechamento_semana_detail', data_inicio=data_inicio.isoformat())
        except Exception as e:
            messages.error(request, f'Erro ao excluir fechamento: {str(e)}')
            return redirect('funcionarios:fechamento_detail', pk=pk)
    
    # GET - mostrar página de confirmação
    context = {
        'fechamento': fechamento,
        'title': 'Confirmar Exclusão'
    }
    return render(request, 'funcionarios/fechamento_delete_confirm.html', context)


@login_required
def fechamento_semana_delete(request, data_inicio):
    """Deleta todos os fechamentos de uma semana com confirmação"""
    try:
        dt_inicio = datetime.date.fromisoformat(data_inicio)
    except ValueError:
        messages.error(request, 'Data inválida.')
        return redirect('funcionarios:fechamento_list')
    
    # Buscar todos os fechamentos daquela semana
    fechamentos = FechamentoSemanal.objects.filter(data_inicio=dt_inicio).select_related('funcionario')
    
    if not fechamentos.exists():
        messages.warning(request, 'Nenhum fechamento encontrado para esta semana.')
        return redirect('funcionarios:fechamento_list')
    
    if request.method == 'POST':
        # Verificar se há fechamentos pagos (proteção)
        fechamentos_pagos = fechamentos.filter(status='pago')
        
        if fechamentos_pagos.exists():
            qtd_pagos = fechamentos_pagos.count()
            nomes_pagos = ', '.join([f.funcionario.nome_completo for f in fechamentos_pagos])
            messages.error(
                request,
                f'Não é possível excluir esta semana. {qtd_pagos} fechamento(s) já foi/foram pago(s): {nomes_pagos}'
            )
            return redirect('funcionarios:fechamento_semana_detail', data_inicio=data_inicio)
        
        qtd_deletados = fechamentos.count()
        total_valor = fechamentos.aggregate(Sum('total_valor'))['total_valor__sum'] or Decimal('0.00')
        
        try:
            fechamentos.delete()
            total_valor_str = brl(total_valor)
            messages.success(
                request,
                f'✅ Semana de {dt_inicio.strftime("%d/%m/%Y")} excluída com sucesso! '
                f'{qtd_deletados} fechamento(s) deletado(s) ({total_valor_str}).'
            )
            return redirect('funcionarios:fechamento_list')
        except Exception as e:
            messages.error(request, f'Erro ao excluir semana: {str(e)}')
            return redirect('funcionarios:fechamento_semana_detail', data_inicio=data_inicio)
    
    # GET - mostrar página de confirmação
    totais = fechamentos.aggregate(
        total_funcionarios=Count('id'),
        total_dias=Sum('total_dias'),
        total_valor=Sum('total_valor'),
        qtd_fechados=Count('id', filter=Q(status='fechado')),
        qtd_pagos=Count('id', filter=Q(status='pago')),
    )
    
    context = {
        'data_inicio': dt_inicio,
        'data_fim': fechamentos.first().data_fim,
        'fechamentos': fechamentos,
        'totais': totais,
        'title': 'Confirmar Exclusão da Semana'
    }
    return render(request, 'funcionarios/fechamento_semana_delete_confirm.html', context)


@login_required
def fechamento_auto(request):
    """Gera fechamentos automaticamente para todos os funcionários ativos num período"""
    if request.method == 'POST':
        data_inicio_str = request.POST.get('data_inicio')
        data_fim_str = request.POST.get('data_fim')
        if not data_inicio_str or not data_fim_str:
            messages.error(request, 'Informe a data de início e a data de fim do período.')
            return redirect('funcionarios:fechamento_auto')

        try:
            data_inicio = datetime.date.fromisoformat(data_inicio_str)
            data_fim = datetime.date.fromisoformat(data_fim_str)
        except ValueError:
            messages.error(request, 'Data inválida.')
            return redirect('funcionarios:fechamento_auto')

        if data_fim < data_inicio:
            messages.error(request, 'Data fim não pode ser anterior à data início.')
            return redirect('funcionarios:fechamento_auto')

        # Verifica se já existe fechamento para este período
        fechamentos_existentes = FechamentoSemanal.objects.filter(
            data_inicio=data_inicio, data_fim=data_fim
        )
        if fechamentos_existentes.exists():
            qtd = fechamentos_existentes.count()
            messages.warning(
                request,
                f'Já existe(m) {qtd} fechamento(s) para o período de '
                f'{data_inicio.strftime("%d/%m/%Y")} a {data_fim.strftime("%d/%m/%Y")}. '
                f'Não é possível gerar novos fechamentos para este período.'
            )
            return redirect('funcionarios:fechamento_list')

        funcionarios = Funcionario.objects.filter(ativo=True).exclude(funcao='fiscal')
        criados = 0
        existentes = 0

        for func in funcionarios:
            has_ap = ApontamentoFuncionario.objects.filter(
                funcionario=func,
                data__gte=data_inicio,
                data__lte=data_fim
            ).exists()
            if not has_ap:
                continue

            existing = FechamentoSemanal.objects.filter(
                funcionario=func, data_inicio=data_inicio, data_fim=data_fim
            ).first()
            if existing:
                existing.calcular_totais()
                existentes += 1
            else:
                f = FechamentoSemanal(
                    funcionario=func,
                    data_inicio=data_inicio,
                    data_fim=data_fim,
                    status='fechado'
                )
                f.save()
                f.calcular_totais()
                criados += 1

        messages.success(request, f'Fechamento automático concluído: {criados} criados, {existentes} atualizados.')
        return redirect('funcionarios:fechamento_list')

    # GET: show form to select period
    hoje = datetime.date.today()
    # Default: início da semana (segunda)
    inicio_semana = hoje - datetime.timedelta(days=hoje.weekday())
    fim_semana = inicio_semana + datetime.timedelta(days=6)
    context = {
        'title': 'Fechamento Automático',
        'today': hoje,
        'inicio_semana': inicio_semana,
        'fim_semana': fim_semana,
    }
    return render(request, 'funcionarios/fechamento_auto.html', context)


# ==================== VISÕES ESPECIAIS ====================

@login_required
def obra_mao_de_obra(request, pk):
    """Visão de mão de obra por obra: custos por etapa + timeline"""
    obra = get_object_or_404(Obra, pk=pk)
    etapas = Etapa.objects.filter(obra=obra)

    # Custos por etapa
    custo_por_etapa = []
    for etapa in etapas:
        aps = ApontamentoFuncionario.objects.filter(obra=obra, etapa=etapa).exclude(funcionario__funcao='fiscal')
        total = aps.aggregate(
            total_valor=Sum('valor_diaria'),
            total_horas=Sum('horas_trabalhadas'),
            total_dias=Count('data', distinct=True),
        )
        funcionarios_etapa = aps.values(
            'funcionario__nome_completo', 'funcionario__funcao'
        ).annotate(
            dias=Count('data', distinct=True),
            horas=Sum('horas_trabalhadas'),
            valor=Sum('valor_diaria'),
        ).order_by('-dias')

        custo_por_etapa.append({
            'etapa': etapa,
            'total_valor': total['total_valor'] or Decimal('0.00'),
            'total_horas': total['total_horas'] or Decimal('0.0'),
            'total_dias': total['total_dias'] or 0,
            'funcionarios': funcionarios_etapa,
        })

    # Custo total
    custo_total = ApontamentoFuncionario.objects.filter(obra=obra).exclude(funcionario__funcao='fiscal').aggregate(
        total=Sum('valor_diaria')
    )['total'] or Decimal('0.00')

    # Timeline: últimos 30 dias
    hoje = datetime.date.today()
    data_inicio = hoje - datetime.timedelta(days=30)
    timeline_qs = ApontamentoFuncionario.objects.filter(
        obra=obra, data__gte=data_inicio
    ).select_related('funcionario', 'etapa').order_by('data')

    timeline = defaultdict(list)
    for ap in timeline_qs:
        timeline[ap.data].append(ap)

    # Sort timeline
    timeline_sorted = sorted(timeline.items(), key=lambda x: x[0], reverse=True)

    context = {
        'obra': obra,
        'etapas': etapas,
        'custo_por_etapa': custo_por_etapa,
        'custo_total': custo_total,
        'timeline': timeline_sorted,
        'title': f'Mão de Obra - {obra.nome}'
    }
    return render(request, 'funcionarios/obra_mao_de_obra.html', context)


@login_required
def funcionario_historico(request, pk):
    """Histórico semanal/mensal de um funcionário"""
    funcionario = get_object_or_404(Funcionario, pk=pk)

    # Month selector
    mes_str = request.GET.get('mes')
    if mes_str:
        try:
            ano, mes = map(int, mes_str.split('-'))
        except ValueError:
            hoje = datetime.date.today()
            ano, mes = hoje.year, hoje.month
    else:
        hoje = datetime.date.today()
        ano, mes = hoje.year, hoje.month

    # Build calendar data
    cal = calendar.Calendar(firstweekday=0)  # Monday first
    month_days = cal.monthdayscalendar(ano, mes)

    apontamentos_mes = ApontamentoFuncionario.objects.filter(
        funcionario=funcionario,
        data__year=ano,
        data__month=mes
    ).select_related('obra', 'etapa')

    # Map by day
    ap_by_day = {}
    for ap in apontamentos_mes:
        ap_by_day[ap.data.day] = ap

    # Build calendar with data
    calendar_weeks = []
    for week in month_days:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append(None)
            else:
                ap = ap_by_day.get(day)
                week_data.append({
                    'day': day,
                    'date': datetime.date(ano, mes, day),
                    'apontamento': ap,
                })
        calendar_weeks.append(week_data)

    # Stats for the month (one diária per unique date)
    stats = apontamentos_mes.aggregate(
        total_dias=Count('data', distinct=True),
        total_horas=Sum('horas_trabalhadas'),
        total_valor=Sum('valor_diaria'),
        dias_ociosidade=Count('data', filter=Q(houve_ociosidade=True), distinct=True),
        dias_retrabalho=Count('data', filter=Q(houve_retrabalho=True), distinct=True),
    )

    # Obras trabalhadas no mês
    obras_mes = apontamentos_mes.values('obra__nome').annotate(
        dias=Count('data', distinct=True),
        horas=Sum('horas_trabalhadas'),
    ).order_by('-dias')

    # Navigation
    prev_month = datetime.date(ano, mes, 1) - datetime.timedelta(days=1)
    next_month = datetime.date(ano, mes, 1) + datetime.timedelta(days=32)
    next_month = next_month.replace(day=1)

    MESES_PT = [
        '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro',
    ]

    context = {
        'funcionario': funcionario,
        'calendar_weeks': calendar_weeks,
        'stats': stats,
        'obras_mes': obras_mes,
        'ano': ano,
        'mes': mes,
        'mes_nome': MESES_PT[mes],
        'prev_month': f"{prev_month.year}-{prev_month.month:02d}",
        'next_month': f"{next_month.year}-{next_month.month:02d}",
        'current_month': f"{ano}-{mes:02d}",
        'title': f'Histórico - {funcionario.nome_completo}'
    }
    return render(request, 'funcionarios/funcionario_historico.html', context)


@login_required
def funcionario_medias_individuais(request, pk):
    """
    ✅ CORREÇÃO PROBLEMA 3: Mostra médias de rendimento individual de um pedreiro por indicador.
    """
    funcionario = get_object_or_404(Funcionario, pk=pk, funcao='pedreiro')
    
    # Período (padrão: últimos 90 dias)
    hoje = timezone.now().date()
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    if data_inicio and data_fim:
        try:
            data_inicio = datetime.datetime.strptime(data_inicio, '%Y-%m-%d').date()
            data_fim = datetime.datetime.strptime(data_fim, '%Y-%m-%d').date()
        except ValueError:
            data_inicio = hoje - datetime.timedelta(days=90)
            data_fim = hoje
    else:
        data_inicio = hoje - datetime.timedelta(days=90)
        data_fim = hoje
    
    # Buscar produções do pedreiro
    producoes = RegistroProducao.objects.filter(
        funcionario=funcionario,
        data__range=[data_inicio, data_fim]
    )
    
    # Calcular médias por indicador
    from collections import defaultdict
    medias_por_etapa = defaultdict(list)
    
    # Mapear indicadores para etapas
    INDICADOR_ETAPA = {
        'alicerce_percentual': (1, 'Levantar Alicerce', '%'),
        'parede_7fiadas': (1, 'Parede 7 Fiadas', 'blocos'),
        'respaldo_conclusao': (2, 'Respaldo', '%'),
        'laje_conclusao': (2, 'Laje', '%'),
        'platibanda': (2, 'Platibanda', 'blocos'),
        'cobertura_conclusao': (2, 'Cobertura', '%'),
        'reboco_externo': (3, 'Reboco Externo', 'm²'),
        'reboco_interno': (3, 'Reboco Interno', 'm²'),
    }
    
    for indicador_code in producoes.order_by().values_list('indicador', flat=True).distinct():
        prods = producoes.filter(indicador=indicador_code)
        
        if prods.exists():
            # Calcular média CORRETA (evitando o problema de divisão errada)
            total_producao = sum(float(p.quantidade) for p in prods)
            total_dias = prods.values('data').distinct().count()
            
            if total_dias > 0:
                media = total_producao / total_dias
            else:
                media = 0
            
            etapa_num, nome, unidade = INDICADOR_ETAPA.get(
                indicador_code, 
                (0, indicador_code, '')
            )
            
            medias_por_etapa[etapa_num].append({
                'codigo': indicador_code,
                'nome': nome,
                'media': round(media, 2),
                'unidade': unidade,
                'total_producao': round(total_producao, 2),
                'total_dias': total_dias,
            })
    
    # Calcular totais gerais
    total_dias_trabalhados = producoes.values('data').distinct().count()
    total_obras = producoes.values('obra').distinct().count()
    
    context = {
        'funcionario': funcionario,
        'medias_por_etapa': dict(medias_por_etapa),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total_dias_trabalhados': total_dias_trabalhados,
        'total_obras': total_obras,
        'title': f'Médias - {funcionario.nome_completo}'
    }
    
    return render(request, 'funcionarios/funcionario_medias_individuais.html', context)


# ==================== APIs ====================

@login_required
def check_fechamento_api(request):
    """API JSON para verificar se já existe fechamento para uma semana."""
    data_inicio_str = request.GET.get('data_inicio', '')
    funcionario_id = request.GET.get('funcionario', '')

    if not data_inicio_str:
        return JsonResponse({'error': 'Parâmetro data_inicio obrigatório'}, status=400)

    try:
        di = datetime.date.fromisoformat(data_inicio_str)
    except ValueError:
        return JsonResponse({'error': 'data_inicio inválida'}, status=400)

    qs = FechamentoSemanal.objects.filter(data_inicio=di)
    if funcionario_id:
        qs_func = qs.filter(funcionario_id=funcionario_id)
        exists_func = qs_func.exists()
    else:
        exists_func = False

    exists_any = qs.exists()
    count = qs.count()

    funcionarios_list = list(
        qs.select_related('funcionario')
        .values_list('funcionario__nome_completo', flat=True)
        .order_by('funcionario__nome_completo')[:10]
    )

    di_fim = di + datetime.timedelta(days=6)
    return JsonResponse({
        'exists': exists_any,
        'exists_funcionario': exists_func,
        'count': count,
        'funcionarios': funcionarios_list,
        'semana': f'{di.strftime("%d/%m/%Y")} a {di_fim.strftime("%d/%m/%Y")}',
    })


@login_required
def apontamentos_api(request):
    """API JSON para retornar apontamentos de um funcionário em um intervalo de datas."""
    funcionario_id = request.GET.get('funcionario')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    if not funcionario_id or not data_inicio:
        return JsonResponse({'error': 'Parâmetros insuficientes'}, status=400)

    try:
        di = datetime.date.fromisoformat(data_inicio)
    except Exception:
        return JsonResponse({'error': 'data_inicio inválida'}, status=400)

    if data_fim:
        try:
            df = datetime.date.fromisoformat(data_fim)
        except Exception:
            return JsonResponse({'error': 'data_fim inválida'}, status=400)
    else:
        df = di + datetime.timedelta(days=6)

    qs = ApontamentoFuncionario.objects.filter(
        funcionario_id=funcionario_id,
        data__gte=di,
        data__lte=df
    ).select_related('obra', 'etapa').order_by('data')

    apontamentos = []
    total_valor = Decimal('0.00')
    total_horas = Decimal('0.0')
    seen_dates = set()
    for a in qs:
        apontamentos.append({
            'id': a.id,
            'obra': a.obra.nome if a.obra else None,
            'etapa': str(a.etapa) if a.etapa else None,
            'data': a.data.isoformat(),
            'horas_trabalhadas': str(a.horas_trabalhadas),
            'clima': a.clima,
            'houve_ociosidade': a.houve_ociosidade,
            'houve_retrabalho': a.houve_retrabalho,
            'valor_diaria': str(a.valor_diaria),
            'created_at': a.created_at.isoformat(),
        })
        total_valor += a.valor_diaria or Decimal('0.00')
        total_horas += a.horas_trabalhadas or Decimal('0.0')
        seen_dates.add(a.data)

    result = {
        'apontamentos': apontamentos,
        'totais': {
            'dias': len(seen_dates),
            'horas': str(total_horas),
            'valor': f"{total_valor:.2f}"
        }
    }
    return JsonResponse(result)


@login_required
@require_GET
def obras_autocomplete_api(request):
    """Autocomplete API for Obra by name. Returns limited list to avoid heavy queries.

    Query params:
    - q: partial name (required, min 2 chars recommended)
    - limit: max results (optional, default 10, max 50)
    """
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 1:
        return JsonResponse({'results': []})

    try:
        limit = int(request.GET.get('limit', 10))
    except Exception:
        limit = 10
    limit = max(1, min(limit, 50))

    # Basic search: name contains q (case-insensitive). Use only active, planning or in_progress
    obras_qs = Obra.objects.filter(
        ativo=True,
        status__in=['planejamento', 'em_andamento'],
        nome__icontains=q
    ).order_by('nome')[:limit]

    results = []
    for o in obras_qs:
        results.append({'id': o.pk, 'text': o.nome})

    return JsonResponse({'results': results})


@login_required
def etapas_por_obra_api(request):
    """API para retornar etapas de uma obra (para preencher select dinamicamente)"""
    obra_id = request.GET.get('obra_id')
    if not obra_id:
        return JsonResponse({'etapas': []})

    etapas = Etapa.objects.filter(
        obra_id=obra_id,
        obra__ativo=True,
        obra__status__in=['planejamento', 'em_andamento'],
        status='em_andamento',
    ).order_by('numero_etapa')
    data = [{'id': e.id, 'label': e.get_numero_etapa_display()} for e in etapas]
    return JsonResponse({'etapas': data})


@login_required
def itens_etapa_api(request):
    """API para retornar os itens/campos de uma etapa específica com valores atuais."""
    etapa_id = request.GET.get('etapa_id')
    if not etapa_id:
        return JsonResponse({'items': []})
    try:
        etapa = Etapa.objects.get(
            pk=etapa_id,
            obra__ativo=True,
            obra__status__in=['planejamento', 'em_andamento'],
            status='em_andamento',
        )
    except Etapa.DoesNotExist:
        return JsonResponse({'items': []})

    items = _get_etapa_items(etapa)
    return JsonResponse({
        'items': items,
        'etapa_id': etapa.id,
        'etapa_numero': etapa.numero_etapa,
        'etapa_label': etapa.get_numero_etapa_display(),
    })


@login_required
def itens_obra_api(request):
    """API para retornar TODAS as etapas com seus itens para uma obra."""
    obra_id = request.GET.get('obra_id')
    if not obra_id:
        return JsonResponse({'etapas': []})

    etapas = Etapa.objects.filter(obra_id=obra_id).order_by('numero_etapa')
    result = []
    for etapa in etapas:
        items = _get_etapa_items(etapa)
        result.append({
            'id': etapa.id,
            'numero': etapa.numero_etapa,
            'label': etapa.get_numero_etapa_display(),
            'concluida': etapa.concluida,
            'items': items,
        })
    return JsonResponse({'etapas': result})


# ================ APONTAMENTO EM LOTE ================

@login_required
@transaction.atomic
def apontamento_lote_create(request):
    """Cria apontamento em lote para múltiplos funcionários"""
    
    if request.method == 'POST':
        form_lote = ApontamentoDiarioLoteForm(request.POST)
        
        if form_lote.is_valid():
            # Validar funcionários
            funcionarios_ids = request.POST.getlist('funcionario')
            horas_trabalhadas_list = request.POST.getlist('horas_trabalhadas')
            
            # Remover valores vazios
            funcionarios_ids = [f for f in funcionarios_ids if f]
            
            if not funcionarios_ids:
                messages.error(request, '❌ Adicione pelo menos 1 funcionário!')
                context = {
                    'form': form_lote,
                    'funcionarios': Funcionario.objects.filter(ativo=True).order_by('nome_completo'),
                    'title': 'Apontamento Diário em Lote'
                }
                return render(request, 'funcionarios/apontamento_lote_form.html', context)
            
            # Fluxo multi-etapas: processa todos os rascunhos salvos no frontend.
            rascunhos_json = request.POST.get('rascunhos_etapas_json', '').strip()
            if rascunhos_json:
                try:
                    rascunhos = json.loads(rascunhos_json)
                except (TypeError, ValueError, json.JSONDecodeError):
                    rascunhos = []

                if isinstance(rascunhos, list) and rascunhos:
                    base_data = {
                        'obra': form_lote.cleaned_data.get('obra'),
                        'data': form_lote.cleaned_data.get('data'),
                        'clima': form_lote.cleaned_data.get('clima'),
                        'houve_ociosidade': form_lote.cleaned_data.get('houve_ociosidade', False),
                        'observacao_ociosidade': form_lote.cleaned_data.get('observacao_ociosidade'),
                        'houve_retrabalho': form_lote.cleaned_data.get('houve_retrabalho', False),
                        'motivo_retrabalho': form_lote.cleaned_data.get('motivo_retrabalho'),
                        'possui_placa': form_lote.cleaned_data.get('possui_placa', False),
                        'observacoes': form_lote.cleaned_data.get('observacoes'),
                    }
                    fotos_uploaded = request.FILES.getlist('fotos')
                    lotes_criados = []
                    etapas_invalidas = 0

                    for idx, rascunho in enumerate(reversed(rascunhos)):
                        if not isinstance(rascunho, dict):
                            continue
                        etapa_id = rascunho.get('etapa_id')
                        if not etapa_id:
                            continue

                        etapa_obj = Etapa.objects.filter(
                            pk=etapa_id,
                            obra=base_data['obra'],
                            status='em_andamento'
                        ).first()
                        if not etapa_obj:
                            etapas_invalidas += 1
                            continue

                        lote, funcionarios_criados, apontamentos_criados = _criar_lote_por_payload(
                            base_data=base_data,
                            etapa=etapa_obj,
                            request=request,
                            funcionarios_ids=funcionarios_ids,
                            horas_trabalhadas_list=horas_trabalhadas_list,
                            campos_payload=rascunho.get('campos') or {},
                            fotos=fotos_uploaded if idx == 0 else []
                        )
                        if lote:
                            lotes_criados.append((lote, funcionarios_criados, apontamentos_criados))

                    if lotes_criados:
                        total_etapas = len(lotes_criados)
                        total_funcionarios = sum(item[1] for item in lotes_criados)
                        total_apontamentos = sum(item[2] for item in lotes_criados)

                        messages.success(
                            request,
                            f'✅ {total_etapas} etapa(s) salva(s) com sucesso! '
                            f'{total_funcionarios} registro(s) de funcionários e {total_apontamentos} apontamento(s) individual(is) gerado(s).'
                        )
                        if etapas_invalidas:
                            messages.warning(request, f'⚠️ {etapas_invalidas} etapa(s) não puderam ser salvas.')
                        if fotos_uploaded:
                            messages.info(request, f'📷 {len(fotos_uploaded)} foto(s) anexada(s)!')
                        return redirect('funcionarios:apontamento_list')

                    messages.error(request, '❌ Nenhuma etapa salva no rascunho pôde ser processada.')
                    context = {
                        'form': form_lote,
                        'funcionarios': Funcionario.objects.filter(ativo=True).order_by('nome_completo'),
                        'title': 'Apontamento Diário em Lote'
                    }
                    return render(request, 'funcionarios/apontamento_lote_form.html', context)

            # Salvar lote
            lote = form_lote.save(commit=False)
            lote.criado_por = request.user
            lote.save()
            
            # Criar registros FuncionarioLote
            funcionarios_criados = 0
            for i, func_id in enumerate(funcionarios_ids):
                try:
                    funcionario = Funcionario.objects.get(pk=func_id, ativo=True)
                    horas = Decimal(horas_trabalhadas_list[i]) if i < len(horas_trabalhadas_list) else Decimal('8.0')
                    if funcionario.funcao == 'fiscal':
                        horas = Decimal('0.0')
                    elif horas <= Decimal('0.0'):
                        horas = Decimal('8.0')
                    
                    FuncionarioLote.objects.create(
                        lote=lote,
                        funcionario=funcionario,
                        horas_trabalhadas=horas
                    )
                    funcionarios_criados += 1
                except (Funcionario.DoesNotExist, ValueError, InvalidOperation):
                    continue
            
            if funcionarios_criados == 0:
                lote.delete()
                messages.error(request, '❌ Nenhum funcionário válido foi adicionado!')
                context = {
                    'form': form_lote,
                    'funcionarios': Funcionario.objects.filter(ativo=True).order_by('nome_completo'),
                    'title': 'Apontamento Diário em Lote'
                }
                return render(request, 'funcionarios/apontamento_lote_form.html', context)
            
            # ========== PROCESSAR CAMPOS DA ETAPA ==========
            etapa = lote.etapa
            campos_atualizados = []
            valores_producao = []  # Armazena valores numéricos de produção dos campos
            valores_producao_dia = {}  # NOVO: valores adicionados NESTE dia (não acumulados)
            
            if etapa:
                numero_etapa = etapa.numero_etapa
                
                # Buscar ou criar o objeto de detalhes da etapa
                detalhes = None
                try:
                    if numero_etapa == 1:
                        detalhes, created = Etapa1Fundacao.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 2:
                        detalhes, created = Etapa2Estrutura.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 3:
                        detalhes, created = Etapa3Instalacoes.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 4:
                        detalhes, created = Etapa4Acabamentos.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 5:
                        detalhes, created = Etapa5Finalizacao.objects.get_or_create(etapa=etapa)
                except Exception as e:
                    messages.warning(request, f'⚠️ Erro ao criar detalhes da etapa: {str(e)}')
                
                # Atualizar campos com valores do POST
                if detalhes:
                    for key, value in request.POST.items():
                        if key.startswith('campo_'):
                            campo_nome = key.replace('campo_', '')
                            
                            # ✅ CORREÇÃO PROBLEMA 1: Só processar se valor foi realmente informado
                            # Ignorar campos vazios, zero ou espaços em branco
                            if not value or value.strip() == '' or value.strip() == '0' or value.strip() == '0.00':
                                continue  # Pula este campo, não cria registro!
                            
                            # Verificar se o campo existe no model
                            if hasattr(detalhes, campo_nome):
                                try:
                                    field = detalhes._meta.get_field(campo_nome)
                                    valor_anterior = getattr(detalhes, campo_nome, None)
                                    
                                    if field.get_internal_type() == 'BooleanField':
                                        # Checkbox
                                        novo_valor = value == 'on'
                                        if valor_anterior != novo_valor:
                                            setattr(detalhes, campo_nome, novo_valor)
                                            campos_atualizados.append(f"{field.verbose_name}: {'✓' if novo_valor else '✗'}")
                                    
                                    elif field.get_internal_type() == 'DecimalField':
                                        # Decimal - INCREMENTAR valores de produção (blocos, m², etc)
                                        if value:
                                            novo_valor = Decimal(value)
                                            # Capturar valores numéricos de produção (sempre que > 0)
                                            if novo_valor > 0:
                                                valores_producao.append((field.verbose_name, novo_valor))
                                                # GUARDAR valor do DIA (não acumulado) para RegistroProducao
                                                valores_producao_dia[campo_nome] = novo_valor
                                            
                                            # INCREMENTAR valor anterior ao invés de substituir
                                            valor_anterior_decimal = valor_anterior if valor_anterior else Decimal('0.00')
                                            valor_final = valor_anterior_decimal + novo_valor
                                            
                                            # Verificar se campo tem limite máximo (ex: percentuais até 100%)
                                            tem_max_100 = any(
                                                hasattr(v, 'limit_value') and v.limit_value == Decimal('100.00')
                                                for v in field.validators
                                            )
                                            
                                            if tem_max_100 and valor_final > Decimal('100.00'):
                                                # Não permitir ultrapassar 100%
                                                diferenca = Decimal('100.00') - valor_anterior_decimal
                                                if diferenca > 0:
                                                    valor_final = Decimal('100.00')
                                                    campos_atualizados.append(f"{field.verbose_name}: +{diferenca} (total: 100.00% - LIMITE ATINGIDO)")
                                                    messages.warning(request, f'⚠️ {field.verbose_name} atingiu o limite de 100%. Não é possível adicionar mais.')
                                                else:
                                                    messages.error(request, f'❌ {field.verbose_name} já está em 100%. Não é possível adicionar mais.')
                                                    continue
                                            else:
                                                campos_atualizados.append(f"{field.verbose_name}: +{novo_valor} (total: {valor_final})")
                                            
                                            setattr(detalhes, campo_nome, valor_final)
                                    
                                    elif field.get_internal_type() in ['IntegerField', 'PositiveIntegerField']:
                                        # Integer - INCREMENTAR valores de produção (blocos, etc)
                                        if value:
                                            novo_valor = int(value)
                                            # Capturar valores numéricos de produção (sempre que > 0)
                                            if novo_valor > 0:
                                                valores_producao.append((field.verbose_name, Decimal(str(novo_valor))))
                                                # GUARDAR valor do DIA (não acumulado) para RegistroProducao
                                                valores_producao_dia[campo_nome] = novo_valor
                                            
                                            # INCREMENTAR valor anterior ao invés de substituir
                                            valor_anterior_int = valor_anterior if valor_anterior else 0
                                            valor_final = valor_anterior_int + novo_valor
                                            
                                            setattr(detalhes, campo_nome, valor_final)
                                            campos_atualizados.append(f"{field.verbose_name}: +{novo_valor} (total: {valor_final})")
                                    
                                    elif field.get_internal_type() == 'DateField':
                                        # Date
                                        if value:
                                            novo_valor = datetime.datetime.strptime(value, '%Y-%m-%d').date()
                                            if valor_anterior != novo_valor:
                                                setattr(detalhes, campo_nome, novo_valor)
                                                campos_atualizados.append(f"{field.verbose_name}: {novo_valor.strftime('%d/%m/%Y')}")
                                
                                except Exception as e:
                                    messages.warning(request, f'⚠️ Erro ao processar campo {campo_nome}: {str(e)}')
                                    continue
                    
                    # Salvar atualizações
                    if campos_atualizados:
                        detalhes.save()
                        
                        # Registrar no histórico
                        try:
                            EtapaHistorico.objects.create(
                                etapa=etapa,
                                origem='Apontamento em Lote',
                                descricao=f"📝 Campos atualizados via apontamento em lote:\n" + "\n".join([f"  • {c}" for c in campos_atualizados]),
                                usuario=request.user
                            )
                        except Exception:
                            pass
                        
                        messages.info(request, f'📊 {len(campos_atualizados)} campo(s) da etapa atualizado(s)')
                    
                    # ========== CALCULAR PRODUÇÃO TOTAL DOS CAMPOS ==========
                    # Se não foi preenchido producao_total manualmente, usar os campos da etapa
                    producao_total_dia, unidade_dia = _calcular_producao_lote_por_campos(valores_producao_dia)
                    lote.producao_total = producao_total_dia
                    lote.unidade_medida = unidade_dia
                    lote.save(update_fields=['producao_total', 'unidade_medida'])
            
            # ========== FIM PROCESSAMENTO CAMPOS ==========
            
            # Recarregar lote para pegar valores atualizados
            lote.refresh_from_db()
            
            # GUARDAR valores do dia no lote (temporariamente) para usar em _criar_registro_producao
            lote._valores_dia = valores_producao_dia
            
            # Debug: verificar quantidade de pedreiros
            total_funcionarios = lote.funcionarios.count()
            pedreiros_count = sum(1 for f in lote.funcionarios.all() if f.funcionario.funcao == 'pedreiro')
            
            # Gerar apontamentos individuais
            apontamentos_criados = lote.gerar_apontamentos_individuais()
            
            # Mensagem com detalhes da divisão
            if lote.producao_total and lote.producao_total > 0 and pedreiros_count > 0:
                producao_por_pedreiro = (lote.producao_total / Decimal(pedreiros_count)).quantize(Decimal('0.01'))
                unidade_map = {'blocos': 'blocos', 'm2': 'm²', 'percentual': '%'}
                unidade = unidade_map.get(lote.unidade_medida, lote.unidade_medida)
                messages.success(request, f'📐 Divisão: {lote.producao_total} {unidade} ÷ {pedreiros_count} pedreiro(s) = {producao_por_pedreiro} {unidade}/pedreiro')
            
            if apontamentos_criados:
                messages.success(request, f'✅ {apontamentos_criados} apontamento(s) individual(is) criado(s)!')
            
            # ========== PROCESSAR FOTOS ==========
            fotos_uploaded = request.FILES.getlist('fotos')
            for foto in fotos_uploaded:
                FotoApontamento.objects.create(
                    apontamento_lote=lote,
                    obra=lote.obra,
                    etapa=lote.etapa,
                    data_foto=lote.data,
                    foto=foto
                )
            
            if fotos_uploaded:
                messages.info(request, f'📷 {len(fotos_uploaded)} foto(s) anexada(s)!')
            # ========================================
            
            etapa_label = lote.etapa.get_numero_etapa_display() if lote.etapa else 'Sem Etapa'
            msg = (
                f'✅ Apontamento da {etapa_label} criado com sucesso! '
                f'{funcionarios_criados} funcionário(s) registrado(s).'
            )
            
            messages.success(request, msg)
            return redirect('funcionarios:apontamento_list')
    
    else:
        # Se vier com ?obra=ID, pré-carregar o estado de 'possui_placa' baseado
        # no último apontamento desta obra (requisito: manter marcado entre apontamentos).
        initial = {}
        obra_id = request.GET.get('obra') or request.GET.get('obra_id')
        if obra_id:
            try:
                ultimo_individual = ApontamentoFuncionario.objects.filter(
                    obra_id=obra_id
                ).order_by('-data', '-created_at').first()
                ultimo_lote = ApontamentoDiarioLote.objects.filter(
                    obra_id=obra_id
                ).order_by('-data', '-created_at').first()

                possui_placa = False
                if ultimo_individual and ultimo_lote:
                    if ultimo_lote.data > ultimo_individual.data:
                        possui_placa = getattr(ultimo_lote, 'possui_placa', False)
                    elif ultimo_individual.data > ultimo_lote.data:
                        possui_placa = getattr(ultimo_individual, 'possui_placa', False)
                    else:
                        # Mesma data: lote tem prioridade
                        possui_placa = getattr(ultimo_lote, 'possui_placa', False)
                elif ultimo_individual:
                    possui_placa = getattr(ultimo_individual, 'possui_placa', False)
                elif ultimo_lote:
                    possui_placa = getattr(ultimo_lote, 'possui_placa', False)

                if possui_placa:
                    initial['possui_placa'] = True
            except Exception:
                pass

        form_lote = ApontamentoDiarioLoteForm(initial=initial) if initial else ApontamentoDiarioLoteForm()
    
    # Buscar funcionários ativos
    funcionarios = Funcionario.objects.filter(ativo=True).order_by('nome_completo')
    funcionarios_json = json.dumps([
        {
            'id': f.id,
            'nome_completo': f.nome_completo,
            'funcao': f.funcao,
            'funcao_display': f.get_funcao_display()
        } for f in funcionarios
    ])
    
    context = {
        'form': form_lote,
        'funcionarios': funcionarios,
        'funcionarios_json': funcionarios_json,
        'title': 'Apontamento Diário em Lote'
    }
    
    return render(request, 'funcionarios/apontamento_lote_form.html', context)


@login_required
def apontamento_lote_list(request):
    """Lista apontamentos em lote com agrupamento visual por data/obra/etapa"""
    lotes = ApontamentoDiarioLote.objects.select_related(
        'obra', 'etapa', 'criado_por'
    ).prefetch_related('funcionarios__funcionario').order_by('-data', '-created_at')
    
    # Filtros
    obra_id = request.GET.get('obra')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    if obra_id:
        lotes = lotes.filter(obra_id=obra_id)
    if data_inicio:
        try:
            lotes = lotes.filter(data__gte=datetime.datetime.strptime(data_inicio, '%Y-%m-%d').date())
        except ValueError:
            pass
    if data_fim:
        try:
            lotes = lotes.filter(data__lte=datetime.datetime.strptime(data_fim, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    # Agrupar por data + obra + etapa para destacar visualmente
    cores = ['table-primary', 'table-success', 'table-warning', 'table-info', 'table-light']
    grupo_cores = {}
    idx_cor = 0
    lotes_list = list(lotes)
    lote_cor_map = {}
    
    for lote in lotes_list:
        chave = f"{lote.data}_{lote.obra_id}_{lote.etapa_id if lote.etapa else 'sem_etapa'}"
        if chave not in grupo_cores:
            grupo_cores[chave] = cores[idx_cor % len(cores)]
            idx_cor += 1
        lote_cor_map[lote.pk] = grupo_cores[chave]
    
    # Pagination
    paginator = Paginator(lotes, 20)
    page = request.GET.get('page')
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'lotes': page_obj,
        'lote_cor_map': lote_cor_map,
        'title': 'Apontamentos em Lote',
        'obras': Obra.objects.filter(ativo=True).order_by('nome'),
    }
    
    return render(request, 'funcionarios/apontamento_lote_list.html', context)


@login_required
def api_campos_etapa(request):
    """
    Retorna os campos disponíveis para uma etapa específica.
    Usado para carregar dinamicamente os campos no formulário de apontamento.
    """
    etapa_id = request.GET.get('etapa_id')
    
    if not etapa_id:
        return JsonResponse({'error': 'etapa_id não informado'}, status=400)
    
    try:
        etapa = Etapa.objects.get(pk=etapa_id)
    except Etapa.DoesNotExist:
        return JsonResponse({'error': 'Etapa não encontrada'}, status=404)
    
    # Determinar qual model de detalhe usar baseado no número da etapa
    numero_etapa = etapa.numero_etapa
    
    # Buscar valores atuais do model de detalhes
    valores_atuais = {}
    try:
        if numero_etapa == 1:
            detalhes = etapa.fundacao
        elif numero_etapa == 2:
            detalhes = etapa.estrutura
        elif numero_etapa == 3:
            detalhes = etapa.instalacoes
        elif numero_etapa == 4:
            detalhes = etapa.acabamentos
        elif numero_etapa == 5:
            detalhes = etapa.finalizacao
        else:
            detalhes = None
        
        if detalhes:
            for campo in detalhes._meta.get_fields():
                if not campo.auto_created and campo.name != 'etapa' and campo.name != 'id':
                    valor = getattr(detalhes, campo.name, None)
                    if valor is not None:
                        if isinstance(valor, Decimal):
                            valores_atuais[campo.name] = str(valor)
                        elif isinstance(valor, datetime.date):
                            valores_atuais[campo.name] = valor.isoformat()
                        else:
                            valores_atuais[campo.name] = valor
    except Exception:
        pass  # Se não existe detalhes ainda, valores_atuais fica vazio
    
    campos = []
    
    if numero_etapa == 1:
        # Etapa 1 - Fundação
        campos = [
            {
                'nome': 'limpeza_terreno',
                'label': 'Limpeza do Terreno',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('limpeza_terreno', False)
            },
            {
                'nome': 'instalacao_energia_agua',
                'label': 'Instalação de Padrão de Energia e Cavalete d\'Água',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('instalacao_energia_agua', False)
            },
            {
                'nome': 'marcacao_escavacao_inicio',
                'label': 'Marcação e Escavação',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('marcacao_escavacao_inicio', '')
            },
            {
                'nome': 'marcacao_escavacao_conclusao',
                'label': 'Marcação e Escavação',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('marcacao_escavacao_conclusao', '')
            },
            {
                'nome': 'locacao_ferragem_inicio',
                'label': 'Locação de Ferragem e Concretagem',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('locacao_ferragem_inicio', '')
            },
            {
                'nome': 'locacao_ferragem_conclusao',
                'label': 'Locação de Ferragem e Concretagem',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('locacao_ferragem_conclusao', '')
            },
            {
                'nome': 'levantar_alicerce_percentual',
                'label': 'Levantar Alicerce, Reboco e Impermeabilizar',
                'tipo': 'number',
                'unidade': '%',
                'min': 0,
                'max': 100,
                'step': '0.01',
                'help_text': 'Quanto foi executado HOJE (será somado ao total)',
                'valor_atual': '0',  # Campo inicia em zero para evitar soma acidental
                'valor_atual_display': str(valores_atuais.get('levantar_alicerce_percentual', '0.00')),
                'bloqueado': Decimal(str(valores_atuais.get('levantar_alicerce_percentual', '0.00'))) >= Decimal('100.00'),
                'aviso': '✅ 100% concluído!' if Decimal(str(valores_atuais.get('levantar_alicerce_percentual', '0.00'))) >= Decimal('100.00') else None
            },
            {
                'nome': 'aterro_contrapiso_inicio',
                'label': 'Aterrar e Fazer Contra Piso',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('aterro_contrapiso_inicio', '')
            },
            {
                'nome': 'aterro_contrapiso_conclusao',
                'label': 'Aterrar e Fazer Contra Piso',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('aterro_contrapiso_conclusao', '')
            },
            {
                'nome': 'parede_7fiadas_blocos',
                'label': 'Parede - 7 Fiadas',
                'tipo': 'number',
                'unidade': 'blocos',
                'min': 0,
                'step': '1',
                'help_text': 'Quantidade de blocos assentados HOJE (será somado ao total)',
                'valor_atual': '0',  # Campo inicia em zero para evitar soma acidental
                'valor_atual_display': str(valores_atuais.get('parede_7fiadas_blocos', '0'))
            },
            {
                'nome': 'fiadas_respaldo_inicio',
                'label': '8 Fiadas até Respaldo',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('fiadas_respaldo_inicio', '')
            },
            {
                'nome': 'fiadas_respaldo_conclusao',
                'label': '8 Fiadas até Respaldo',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('fiadas_respaldo_conclusao', '')
            },
        ]
    
    elif numero_etapa == 2:
        # Etapa 2 - Estrutura
        campos = [
            {
                'nome': 'montagem_laje_inicio',
                'label': 'Montagem da Laje e Concretagem',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('montagem_laje_inicio', '')
            },
            {
                'nome': 'montagem_laje_conclusao',
                'label': 'Montagem da Laje e Concretagem',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('montagem_laje_conclusao', '')
            },
            {
                'nome': 'platibanda_blocos',
                'label': 'Platibanda',
                'tipo': 'number',
                'unidade': 'blocos',
                'min': 0,
                'step': '1',
                'help_text': 'Quantidade de blocos assentados HOJE (será somado ao total)',
                'valor_atual': '0',  # Campo inicia em zero para evitar soma acidental
                'valor_atual_display': str(valores_atuais.get('platibanda_blocos', '0'))
            },
            {
                'nome': 'cobertura_inicio',
                'label': 'Cobertura Completa',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('cobertura_inicio', '')
            },
            {
                'nome': 'cobertura_conclusao',
                'label': 'Cobertura Completa',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('cobertura_conclusao', '')
            },
        ]
    
    elif numero_etapa == 3:
        # Etapa 3 - Instalações
        campos = [
            {
                'nome': 'reboco_externo_m2',
                'label': 'Reboco Externo',
                'tipo': 'number',
                'unidade': 'm²',
                'min': 0,
                'step': '0.01',
                'help_text': 'Metragem executada HOJE em m² (será somado ao total)',
                'valor_atual': '0',  # Campo inicia em zero para evitar soma acidental
                'valor_atual_display': str(valores_atuais.get('reboco_externo_m2', '0.00'))
            },
            {
                'nome': 'reboco_interno_m2',
                'label': 'Reboco Interno',
                'tipo': 'number',
                'unidade': 'm²',
                'min': 0,
                'step': '0.01',
                'help_text': 'Metragem executada HOJE em m² (será somado ao total)',
                'valor_atual': '0',  # Campo inicia em zero para evitar soma acidental
                'valor_atual_display': str(valores_atuais.get('reboco_interno_m2', '0.00'))
            },
            {
                'nome': 'instalacao_portais',
                'label': 'Instalação de Portais',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('instalacao_portais', False)
            },
            {
                'nome': 'agua_fria',
                'label': 'Água Fria',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('agua_fria', False)
            },
            {
                'nome': 'esgoto',
                'label': 'Esgoto',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('esgoto', False)
            },
            {
                'nome': 'fluvial',
                'label': 'Fluvial',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('fluvial', False)
            },
        ]
    
    elif numero_etapa == 4:
        # Etapa 4 - Acabamentos
        campos = [
            {
                'nome': 'portas_janelas',
                'label': 'Portas e Janelas',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('portas_janelas', False)
            },
            {
                'nome': 'pintura_externa_1demao_inicio',
                'label': 'Pintura Externa 1ª Demão',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('pintura_externa_1demao_inicio', '')
            },
            {
                'nome': 'pintura_externa_1demao_conclusao',
                'label': 'Pintura Externa 1ª Demão',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('pintura_externa_1demao_conclusao', '')
            },
            {
                'nome': 'pintura_interna_1demao_inicio',
                'label': 'Pintura Interna 1ª Demão',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('pintura_interna_1demao_inicio', '')
            },
            {
                'nome': 'pintura_interna_1demao_conclusao',
                'label': 'Pintura Interna 1ª Demão',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('pintura_interna_1demao_conclusao', '')
            },
            {
                'nome': 'assentamento_piso_inicio',
                'label': 'Assentamento de Piso',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('assentamento_piso_inicio', '')
            },
            {
                'nome': 'assentamento_piso_conclusao',
                'label': 'Assentamento de Piso',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('assentamento_piso_conclusao', '')
            },
        ]
    
    elif numero_etapa == 5:
        # Etapa 5 - Finalização
        campos = [
            {
                'nome': 'pintura_externa_2demao_inicio',
                'label': 'Pintura Externa 2ª Demão',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('pintura_externa_2demao_inicio', '')
            },
            {
                'nome': 'pintura_externa_2demao_conclusao',
                'label': 'Pintura Externa 2ª Demão',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('pintura_externa_2demao_conclusao', '')
            },
            {
                'nome': 'pintura_interna_2demao_inicio',
                'label': 'Pintura Interna 2ª Demão',
                'tipo': 'date',
                'help_text': 'Data de início',
                'valor_atual': valores_atuais.get('pintura_interna_2demao_inicio', '')
            },
            {
                'nome': 'pintura_interna_2demao_conclusao',
                'label': 'Pintura Interna 2ª Demão',
                'tipo': 'date',
                'help_text': 'Data de conclusão',
                'valor_atual': valores_atuais.get('pintura_interna_2demao_conclusao', '')
            },
            {
                'nome': 'loucas_metais',
                'label': 'Instalação das Louças e Metais',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('loucas_metais', False)
            },
            {
                'nome': 'eletrica',
                'label': 'Elétrica',
                'tipo': 'checkbox',
                'help_text': 'Concluído?',
                'valor_atual': valores_atuais.get('eletrica', False)
            },
        ]
    
    return JsonResponse({
        'etapa_id': etapa_id,
        'etapa_nome': etapa.get_numero_etapa_display(),
        'numero_etapa': numero_etapa,
        'campos': campos
    })


@login_required
def api_obra_possui_placa(request):
    """API para retornar o último valor de possui_placa de uma obra"""
    obra_id = request.GET.get('obra_id')
    if not obra_id:
        return JsonResponse({'possui_placa': False})
    
    try:
        obra_id = int(obra_id)
    except (ValueError, TypeError):
        return JsonResponse({'possui_placa': False})
    
    # Buscar último apontamento (individual ou lote)
    ultimo_individual = ApontamentoFuncionario.objects.filter(
        obra_id=obra_id
    ).order_by('-data', '-created_at').first()
    
    ultimo_lote = ApontamentoDiarioLote.objects.filter(
        obra_id=obra_id
    ).order_by('-data', '-created_at').first()
    
    # Pegar o mais recente — em caso de empate de data, o LOTE tem prioridade
    possui_placa = False
    if ultimo_individual and ultimo_lote:
        if ultimo_lote.data > ultimo_individual.data:
            possui_placa = ultimo_lote.possui_placa
        elif ultimo_individual.data > ultimo_lote.data:
            possui_placa = ultimo_individual.possui_placa
        else:
            # Mesma data: lote é o registro mestre
            possui_placa = ultimo_lote.possui_placa
    elif ultimo_individual:
        possui_placa = ultimo_individual.possui_placa
    elif ultimo_lote:
        possui_placa = ultimo_lote.possui_placa
    
    return JsonResponse({'possui_placa': possui_placa})


@login_required
@require_GET
def api_lote_etapa_contexto(request):
    """Retorna contexto salvo de um lote para a etapa + data selecionadas."""
    etapa_id = request.GET.get('etapa_id')
    data_str = request.GET.get('data')
    obra_id = request.GET.get('obra_id')

    if not etapa_id or not data_str:
        return JsonResponse({
            'encontrado': False,
            'funcionarios': [],
        })

    try:
        data_ref = datetime.date.fromisoformat(data_str)
    except ValueError:
        return JsonResponse({
            'encontrado': False,
            'funcionarios': [],
        })

    lotes = ApontamentoDiarioLote.objects.filter(
        etapa_id=etapa_id,
        data=data_ref,
    )
    if obra_id:
        lotes = lotes.filter(obra_id=obra_id)

    lote = lotes.select_related('obra', 'etapa').prefetch_related(
        'funcionarios__funcionario'
    ).order_by('-created_at').first()

    if not lote:
        return JsonResponse({
            'encontrado': False,
            'funcionarios': [],
            'data': data_ref.isoformat(),
        })

    funcionarios = []
    for item in lote.funcionarios.all():
        funcionarios.append({
            'id': item.funcionario_id,
            'nome_completo': item.funcionario.nome_completo,
            'funcao': item.funcionario.funcao,
            'funcao_display': item.funcionario.get_funcao_display(),
            'horas_trabalhadas': str(item.horas_trabalhadas),
        })

    return JsonResponse({
        'encontrado': True,
        'lote_id': lote.id,
        'data': lote.data.isoformat(),
        'obra_id': lote.obra_id,
        'etapa_id': lote.etapa_id,
        'clima': lote.clima,
        'funcionarios': funcionarios,
    })


# ================ APONTAMENTO EM LOTE - CRUD (EDIÇÃO / EXCLUSÃO / DETALHES) ================


def reverter_producao_etapa(lote):
    """
    Reverte a produção que foi adicionada ao salvar o lote.
    Subtrai os valores dos campos das etapas baseado nos RegistroProducao.
    """
    if not lote.etapa:
        return

    etapa_num = lote.etapa.numero_etapa

    # Buscar detalhes da etapa
    detalhes = None
    try:
        if etapa_num == 1:
            detalhes = Etapa1Fundacao.objects.filter(etapa=lote.etapa).first()
        elif etapa_num == 2:
            detalhes = Etapa2Estrutura.objects.filter(etapa=lote.etapa).first()
        elif etapa_num == 3:
            detalhes = Etapa3Instalacoes.objects.filter(etapa=lote.etapa).first()
        elif etapa_num == 4:
            detalhes = Etapa4Acabamentos.objects.filter(etapa=lote.etapa).first()
        elif etapa_num == 5:
            detalhes = Etapa5Finalizacao.objects.filter(etapa=lote.etapa).first()
    except Exception:
        return

    if not detalhes:
        return

    # Buscar registros de produção deste lote
    prods = RegistroProducao.objects.filter(
        obra=lote.obra,
        data=lote.data,
        etapa=lote.etapa
    )

    # Mapear indicador para campo do model de detalhes
    CAMPO_MAP = {
        'parede_7fiadas': 'parede_7fiadas_blocos',
        'alicerce_percentual': 'levantar_alicerce_percentual',
        'platibanda': 'platibanda_blocos',
        'reboco_externo': 'reboco_externo_m2',
        'reboco_interno': 'reboco_interno_m2',
    }

    # Reverter por indicador
    indicadores = prods.values_list('indicador', flat=True).distinct()
    for indicador in indicadores:
        total_reverter = prods.filter(indicador=indicador).aggregate(
            total=Sum('quantidade')
        )['total'] or Decimal('0')

        campo_nome = CAMPO_MAP.get(indicador)

        if campo_nome and hasattr(detalhes, campo_nome):
            valor_atual = getattr(detalhes, campo_nome)
            if valor_atual is None:
                continue

            novo_valor = Decimal(str(valor_atual)) - total_reverter
            # Garantir que não fique negativo
            novo_valor = max(novo_valor, Decimal('0'))
            setattr(detalhes, campo_nome, novo_valor)

    detalhes.save()


def recalcular_producao_etapa(etapa):
    """
    Reconstroi os campos numericos da etapa a partir dos RegistroProducao remanescentes.
    Isso evita residuos quando um lote e excluido/editado.
    """
    if not etapa:
        return

    detalhes = _obter_detalhes_etapa(etapa)
    if not detalhes:
        return

    # indicador -> campo no model de detalhes da etapa
    campo_map = {
        'parede_7fiadas': 'parede_7fiadas_blocos',
        'alicerce_percentual': 'levantar_alicerce_percentual',
        'platibanda': 'platibanda_blocos',
        'reboco_externo': 'reboco_externo_m2',
        'reboco_interno': 'reboco_interno_m2',
        'respaldo_conclusao': 'respaldo_conclusao',
        'laje_conclusao': 'laje_conclusao',
        'cobertura_conclusao': 'cobertura_conclusao',
    }

    totais = {
        row['indicador']: row['total'] or Decimal('0')
        for row in (
            RegistroProducao.objects
            .filter(etapa=etapa)
            .values('indicador')
            .annotate(total=Sum('quantidade'))
        )
    }

    update_fields = []
    for indicador, campo in campo_map.items():
        if not hasattr(detalhes, campo):
            continue

        field = detalhes._meta.get_field(campo)
        total = totais.get(indicador, Decimal('0'))

        if field.get_internal_type() in ['IntegerField', 'PositiveIntegerField']:
            novo_valor = int(total)
        elif field.get_internal_type() == 'DecimalField':
            scale = Decimal('1').scaleb(-field.decimal_places)
            novo_valor = Decimal(total).quantize(scale)
        else:
            continue

        valor_atual = getattr(detalhes, campo)
        if valor_atual != novo_valor:
            setattr(detalhes, campo, novo_valor)
            update_fields.append(campo)

    if update_fields:
        detalhes.save(update_fields=update_fields)

    # Mantem percentual da obra consistente apos recalculo.
    try:
        etapa.obra.calcular_percentual()
    except Exception:
        pass


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def apontamento_lote_delete(request, pk):
    """
    Exclui apontamento em lote e REVERTE todas as produções nas etapas.
    """
    lote = get_object_or_404(ApontamentoDiarioLote, pk=pk)

    # Guardar dados para histórico ANTES de excluir
    obra = lote.obra
    etapa = lote.etapa
    data = lote.data
    producao_total = lote.producao_total or Decimal('0')
    funcionarios_nomes = [
        f.funcionario.nome_completo for f in lote.funcionarios.select_related('funcionario').all()
    ]

    # PASSO 1: Reverter produção nas etapas (legado)
    reverter_producao_etapa(lote)

    # PASSO 2: Excluir registros de produção
    RegistroProducao.objects.filter(
        obra=obra,
        data=data,
        etapa=etapa
    ).delete()

    # PASSO 3: Excluir apontamentos individuais vinculados
    ApontamentoFuncionario.objects.filter(
        obra=obra,
        data=data,
        etapa=etapa
    ).delete()

    # PASSO 4: Registrar no histórico
    HistoricoAlteracaoEtapa.objects.create(
        obra=obra,
        etapa=etapa,
        tipo_alteracao='exclusao',
        data_referencia=data,
        descricao=(
            f'Apontamento em lote excluído. '
            f'Produção total: {producao_total}. '
            f'Funcionários: {", ".join(funcionarios_nomes)}'
        ),
        usuario=request.user,
        dados_anteriores={
            'producao_total': str(producao_total),
            'funcionarios': funcionarios_nomes,
            'data': data.isoformat(),
        }
    )

    # PASSO 5: Registrar no EtapaHistorico existente
    if etapa:
        try:
            EtapaHistorico.objects.create(
                etapa=etapa,
                origem='Exclusão de Apontamento em Lote',
                descricao=(
                    f'🗑️ APONTAMENTO EM LOTE EXCLUÍDO\n'
                    f'Data: {data.strftime("%d/%m/%Y")}\n'
                    f'Produção revertida: {producao_total}\n'
                    f'Funcionários: {", ".join(funcionarios_nomes)}'
                ),
                usuario=request.user
            )
        except Exception:
            pass

    # PASSO 6: Excluir o lote
    lote.delete()

    # PASSO 7: Recalcular etapa a partir dos registros remanescentes
    recalcular_producao_etapa(etapa)

    messages.success(
        request,
        f'✅ Apontamento excluído com sucesso! '
        f'Produção revertida e relatórios recalculados.'
    )

    return redirect('funcionarios:apontamento_lote_list')


@login_required
@transaction.atomic
def apontamento_lote_edit(request, pk):
    """
    Edita apontamento em lote existente.
    Reverte valores antigos antes de aplicar os novos.
    """
    lote = get_object_or_404(ApontamentoDiarioLote, pk=pk)

    # Guardar valores antigos
    valores_antigos = {
        'producao_total': str(lote.producao_total or 0),
        'funcionarios': [
            {
                'id': fl.funcionario.id,
                'nome': fl.funcionario.nome_completo,
                'horas': str(fl.horas_trabalhadas),
            }
            for fl in lote.funcionarios.select_related('funcionario').all()
        ],
        'clima': lote.clima,
        'observacoes': lote.observacoes or '',
    }

    funcionarios_atuais = lote.funcionarios.select_related('funcionario').all()

    if request.method == 'POST':
        form = ApontamentoDiarioLoteForm(request.POST, instance=lote)

        if form.is_valid():
            # PASSO 1: Reverter produção antiga
            reverter_producao_etapa(lote)

            # PASSO 2: Excluir registros de produção antigos
            RegistroProducao.objects.filter(
                obra=lote.obra,
                data=lote.data,
                etapa=lote.etapa
            ).delete()

            # PASSO 3: Excluir apontamentos individuais antigos
            ApontamentoFuncionario.objects.filter(
                obra=lote.obra,
                data=lote.data,
                etapa=lote.etapa
            ).delete()

            # PASSO 4: Salvar lote atualizado
            lote_atualizado = form.save()

            # PASSO 5: Atualizar funcionários
            lote.funcionarios.all().delete()

            funcionarios_ids = request.POST.getlist('funcionario')
            horas_trabalhadas_list = request.POST.getlist('horas_trabalhadas')
            funcionarios_ids = [f for f in funcionarios_ids if f]

            funcionarios_criados = 0
            for i, func_id in enumerate(funcionarios_ids):
                try:
                    funcionario = Funcionario.objects.get(pk=func_id, ativo=True)
                    horas = Decimal(horas_trabalhadas_list[i]) if i < len(horas_trabalhadas_list) else Decimal('8.0')
                    if funcionario.funcao == 'fiscal':
                        horas = Decimal('0.0')
                    elif horas <= Decimal('0.0'):
                        horas = Decimal('8.0')
                    FuncionarioLote.objects.create(
                        lote=lote,
                        funcionario=funcionario,
                        horas_trabalhadas=horas,
                    )
                    funcionarios_criados += 1
                except (Funcionario.DoesNotExist, ValueError, InvalidOperation):
                    continue

            if funcionarios_criados == 0:
                messages.error(request, '❌ Nenhum funcionário válido foi adicionado!')
                return redirect('funcionarios:apontamento_lote_edit', pk=lote.pk)

            # PASSO 6: Processar campos da etapa (se houver)
            etapa = lote.etapa
            valores_producao_dia = {}

            if etapa:
                numero_etapa = etapa.numero_etapa
                detalhes = None
                try:
                    if numero_etapa == 1:
                        detalhes, _ = Etapa1Fundacao.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 2:
                        detalhes, _ = Etapa2Estrutura.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 3:
                        detalhes, _ = Etapa3Instalacoes.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 4:
                        detalhes, _ = Etapa4Acabamentos.objects.get_or_create(etapa=etapa)
                    elif numero_etapa == 5:
                        detalhes, _ = Etapa5Finalizacao.objects.get_or_create(etapa=etapa)
                except Exception:
                    detalhes = None

                if detalhes:
                    campos_atualizados = False
                    for key, value in request.POST.items():
                        if key.startswith('campo_'):
                            campo_nome = key.replace('campo_', '')
                            if not value or value.strip() == '' or value.strip() == '0' or value.strip() == '0.00':
                                continue

                            if hasattr(detalhes, campo_nome):
                                try:
                                    field = detalhes._meta.get_field(campo_nome)

                                    if field.get_internal_type() == 'BooleanField':
                                        novo_valor = value == 'on'
                                        setattr(detalhes, campo_nome, novo_valor)
                                        campos_atualizados = True

                                    elif field.get_internal_type() == 'DecimalField':
                                        novo_valor = Decimal(value)
                                        if novo_valor > 0:
                                            valores_producao_dia[campo_nome] = novo_valor
                                            valor_anterior = getattr(detalhes, campo_nome) or Decimal('0.00')
                                            setattr(detalhes, campo_nome, valor_anterior + novo_valor)
                                            campos_atualizados = True

                                    elif field.get_internal_type() in ['IntegerField', 'PositiveIntegerField']:
                                        novo_valor = int(value)
                                        if novo_valor > 0:
                                            valores_producao_dia[campo_nome] = Decimal(str(novo_valor))
                                            valor_anterior = getattr(detalhes, campo_nome) or 0
                                            setattr(detalhes, campo_nome, valor_anterior + novo_valor)
                                            campos_atualizados = True

                                    elif field.get_internal_type() == 'DateField':
                                        novo_valor = datetime.datetime.strptime(value, '%Y-%m-%d').date()
                                        setattr(detalhes, campo_nome, novo_valor)
                                        campos_atualizados = True
                                except Exception:
                                    continue

                    if campos_atualizados:
                        detalhes.save()

            # PASSO 7: Recalcular producao do lote com base nos campos da etapa
            producao_total_dia, unidade_dia = _calcular_producao_lote_por_campos(valores_producao_dia)
            lote.producao_total = producao_total_dia
            lote.unidade_medida = unidade_dia
            lote.save(update_fields=['producao_total', 'unidade_medida'])

            # PASSO 8: Gerar nova producao
            lote._valores_dia = valores_producao_dia
            lote.gerar_apontamentos_individuais()

            # PASSO 8.1: Recalcular etapa com base no estado final do banco
            recalcular_producao_etapa(lote.etapa)

            # PASSO 9: Registrar no historico
            detalhes_unidades = []
            for campo_nome, valor in valores_producao_dia.items():
                campo_lower = (campo_nome or '').lower()
                if 'm2' in campo_lower:
                    unidade = 'm2'
                elif 'percentual' in campo_lower:
                    unidade = '%'
                elif 'bloco' in campo_lower or 'fiadas' in campo_lower:
                    unidade = 'blocos'
                else:
                    unidade = 'unidades'
                detalhes_unidades.append(f'{valor} {unidade} ({campo_nome})')

            HistoricoAlteracaoEtapa.objects.create(
                obra=lote.obra,
                etapa=lote.etapa,
                tipo_alteracao='edicao',
                data_referencia=lote.data,
                descricao=(
                    f'Apontamento em lote editado. '
                    f'Producao anterior: {valores_antigos["producao_total"]}. '
                    f'Producao nova (campo escalar): {lote.producao_total or 0}. '
                    f'Campos do dia: {", ".join(detalhes_unidades) if detalhes_unidades else "sem producao numerica"}'
                ),
                usuario=request.user,
                dados_anteriores=valores_antigos,
                dados_novos={
                    'producao_total': str(lote.producao_total or 0),
                    'funcionarios': list(lote.funcionarios.values_list('funcionario_id', flat=True)),
                }
            )
            if etapa:
                try:
                    EtapaHistorico.objects.create(
                        etapa=etapa,
                        origem='Edição de Apontamento em Lote',
                        descricao=(
                            f'✏️ APONTAMENTO EM LOTE EDITADO\n'
                            f'Data: {lote.data.strftime("%d/%m/%Y")}\n'
                            f'Produção: {valores_antigos["producao_total"]} → {lote.producao_total or 0}'
                        ),
                        usuario=request.user
                    )
                except Exception:
                    pass

            messages.success(request, '✅ Apontamento atualizado com sucesso!')
            return redirect('funcionarios:apontamento_lote_detail', pk=lote.pk)
        else:
            messages.error(request, '❌ Corrija os erros no formulário.')
    else:
        form = ApontamentoDiarioLoteForm(instance=lote)

    # Serializar funcionários para JS
    todos_funcionarios = Funcionario.objects.filter(ativo=True).order_by('nome_completo')
    funcionarios_json = json.dumps([
        {
            'id': f.id,
            'nome': f.nome_completo,
            'funcao': f.funcao,
            'funcao_display': f.get_funcao_display(),
        }
        for f in todos_funcionarios
    ])

    context = {
        'form': form,
        'lote': lote,
        'funcionarios_atuais': funcionarios_atuais,
        'funcionarios': todos_funcionarios,
        'funcionarios_json': funcionarios_json,
        'title': f'Editar Apontamento - {lote.data.strftime("%d/%m/%Y")}',
    }

    return render(request, 'funcionarios/apontamento_lote_edit.html', context)


@login_required
def apontamento_lote_detail(request, pk):
    """
    Mostra detalhes completos do apontamento em lote.
    """
    lote = get_object_or_404(ApontamentoDiarioLote, pk=pk)

    funcionarios_lote = lote.funcionarios.select_related('funcionario').all()

    apontamentos = ApontamentoFuncionario.objects.filter(
        obra=lote.obra,
        data=lote.data,
        etapa=lote.etapa,
    ).select_related('funcionario')

    producoes = RegistroProducao.objects.filter(
        obra=lote.obra,
        data=lote.data,
        etapa=lote.etapa,
    ).select_related('funcionario')

    historico = HistoricoAlteracaoEtapa.objects.filter(
        obra=lote.obra,
        etapa=lote.etapa,
        data_referencia=lote.data,
    ).order_by('-created_at')

    detalhes_producao = lote.get_detalhes_producao()

    context = {
        'lote': lote,
        'funcionarios_lote': funcionarios_lote,
        'apontamentos': apontamentos,
        'producoes': producoes,
        'historico': historico,
        'detalhes_producao': detalhes_producao,
        'title': f'Detalhes - Apontamento {lote.data.strftime("%d/%m/%Y")}',
    }

    return render(request, 'funcionarios/apontamento_lote_detail.html', context)
