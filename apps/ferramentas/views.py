from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db import models
from django.http import Http404, HttpResponse
from django.template.loader import render_to_string
from .models import (
    Ferramenta, ConferenciaFerramenta, MovimentacaoFerramenta, 
    ItemConferencia, LocalizacaoFerramenta
)
from .forms import FerramentaForm, MovimentacaoForm, ConferenciaForm, ItemConferenciaForm
from django.contrib import messages
from apps.obras.models import Obra
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import generic
from django.urls import reverse_lazy
from django.views import View
from django.forms import inlineformset_factory
from django.db import transaction
from django.shortcuts import render
from django.core.paginator import Paginator
from django.utils import timezone
from openpyxl import Workbook
import random
from decimal import Decimal
import json


def _filtrar_ferramentas_queryset(request):
    """Aplica os filtros de listagem de ferramentas."""
    qs = Ferramenta.objects.select_related('fornecedor').all()

    codigo = request.GET.get('codigo')
    nome = request.GET.get('nome')
    categoria = request.GET.get('categoria')
    classificacao = request.GET.get('classificacao', '').strip()
    fornecedor = request.GET.get('fornecedor', '').strip()
    status = request.GET.get('status', '').strip()
    busca = request.GET.get('q', '').strip()

    if status == 'ativas':
        qs = qs.filter(ativo=True)
    elif status == 'inativas':
        qs = qs.filter(ativo=False)

    if busca:
        qs = qs.filter(
            models.Q(nome__icontains=busca) |
            models.Q(codigo__icontains=busca)
        )
    if codigo:
        qs = qs.filter(codigo__icontains=codigo)
    if nome:
        qs = qs.filter(nome__icontains=nome)
    if categoria:
        qs = qs.filter(categoria=categoria)
    if classificacao:
        qs = qs.filter(classificacao=classificacao)
    if fornecedor:
        qs = qs.filter(fornecedor_id=fornecedor)

    return qs, busca, categoria, classificacao, fornecedor, status


def _build_ferramenta_relatorio_data(request, paginate=True):
    ids = request.GET.getlist('ids')
    ferramenta_filtro = request.GET.get('ferramenta', '').strip()
    fornecedor_filtro = request.GET.get('fornecedor', '').strip()
    tipo_filtro = request.GET.get('tipo', '').strip()
    origem_filtro = request.GET.get('origem', '').strip()
    destino_filtro = request.GET.get('destino', '').strip()
    data_inicial = request.GET.get('data_inicial', '').strip()
    data_final = request.GET.get('data_final', '').strip()

    if ids:
        base_ferramentas_qs = Ferramenta.objects.select_related('fornecedor').filter(id__in=ids)
    else:
        base_ferramentas_qs, _, _, _, _, _ = _filtrar_ferramentas_queryset(request)

    ferramentas_filtro_opcoes = list(
        base_ferramentas_qs.order_by('nome').values('id', 'codigo', 'nome')
    )

    if ferramenta_filtro:
        base_ferramentas_qs = base_ferramentas_qs.filter(id=ferramenta_filtro)
    if fornecedor_filtro:
        base_ferramentas_qs = base_ferramentas_qs.filter(fornecedor_id=fornecedor_filtro)

    ferramentas_lista = list(base_ferramentas_qs.order_by('nome'))
    ferramenta_ids = [f.id for f in ferramentas_lista]

    total_modelos = len(ferramentas_lista)
    total_unidades = sum((f.quantidade_total or 0) for f in ferramentas_lista)
    total_deposito = sum((f.quantidade_deposito or 0) for f in ferramentas_lista)
    total_em_obra = sum((f.quantidade_em_obras or 0) for f in ferramentas_lista)
    total_manutencao = sum((f.quantidade_manutencao or 0) for f in ferramentas_lista)
    total_perdida = sum((f.quantidade_perdida or 0) for f in ferramentas_lista)

    total_descartada = MovimentacaoFerramenta.objects.filter(
        ferramenta_id__in=ferramenta_ids,
        tipo='descarte'
    ).aggregate(total=models.Sum('quantidade'))['total'] or 0

    obras_distribuicao_qs = (
        LocalizacaoFerramenta.objects
        .filter(ferramenta_id__in=ferramenta_ids, local_tipo='obra', quantidade__gt=0)
        .values('obra_id', 'obra__nome')
        .annotate(total=models.Sum('quantidade'))
        .order_by('obra__nome')
    )

    movimentacoes_qs = (
        MovimentacaoFerramenta.objects
        .filter(ferramenta_id__in=ferramenta_ids)
        .select_related('ferramenta', 'ferramenta__fornecedor', 'responsavel', 'obra_origem', 'obra_destino')
        .order_by('-data_movimentacao', '-id')
    )

    if tipo_filtro:
        movimentacoes_qs = movimentacoes_qs.filter(tipo=tipo_filtro)
    if origem_filtro:
        movimentacoes_qs = movimentacoes_qs.filter(
            models.Q(origem_tipo__icontains=origem_filtro) |
            models.Q(obra_origem__nome__icontains=origem_filtro)
        )
    if destino_filtro:
        movimentacoes_qs = movimentacoes_qs.filter(
            models.Q(destino_tipo__icontains=destino_filtro) |
            models.Q(obra_destino__nome__icontains=destino_filtro)
        )
    if data_inicial:
        movimentacoes_qs = movimentacoes_qs.filter(data_movimentacao__date__gte=data_inicial)
    if data_final:
        movimentacoes_qs = movimentacoes_qs.filter(data_movimentacao__date__lte=data_final)

    def _query_sem(*remove_keys):
        params = request.GET.copy()
        for key in remove_keys:
            params.pop(key, None)
        return params.urlencode()

    if paginate:
        per_page = 10
        ferramentas = Paginator(ferramentas_lista, per_page).get_page(request.GET.get('page_ferramenta', 1))
        obras_distribuicao = Paginator(obras_distribuicao_qs, per_page).get_page(request.GET.get('page_obra', 1))
        movimentacoes = Paginator(movimentacoes_qs, per_page).get_page(request.GET.get('page_historico', 1))
    else:
        ferramentas = ferramentas_lista
        obras_distribuicao = list(obras_distribuicao_qs)
        movimentacoes = list(movimentacoes_qs)

    return {
        'title': 'Relatório de Ferramentas',
        'ferramentas': ferramentas,
        'ferramentas_lista': ferramentas_lista,
        'movimentacoes': movimentacoes,
        'movimentacoes_lista': list(movimentacoes_qs) if paginate else movimentacoes,
        'obras_distribuicao': obras_distribuicao,
        'obras_distribuicao_lista': list(obras_distribuicao_qs) if paginate else obras_distribuicao,
        'tipo_choices': MovimentacaoFerramenta.TIPO_CHOICES,
        'ferramentas_filtro_opcoes': ferramentas_filtro_opcoes,
        'fornecedor_choices': Ferramenta.objects.filter(fornecedor__isnull=False).values_list('fornecedor__id', 'fornecedor__nome').distinct().order_by('fornecedor__nome'),
        'obra_choices': Obra.objects.filter(ativo=True).order_by('nome').values_list('nome', flat=True),
        'ids_selecionados': ids,
        'qs_sem_page_obras': _query_sem('page_obra'),
        'qs_sem_page_ferramenta': _query_sem('page_ferramenta'),
        'qs_sem_page_historico': _query_sem('page_historico'),
        'export_query': _query_sem('page_obra', 'page_ferramenta', 'page_historico', 'export'),
        'total_modelos': total_modelos,
        'total_unidades': total_unidades,
        'total_deposito': total_deposito,
        'total_em_obra': total_em_obra,
        'total_manutencao': total_manutencao,
        'total_perdida': total_perdida,
        'total_descartada': total_descartada,
        'gerado_em': timezone.localtime(),
        'filtros': {
            'q': request.GET.get('q', '').strip(),
            'categoria': request.GET.get('categoria', '').strip(),
            'status': request.GET.get('status', '').strip(),
            'ferramenta': ferramenta_filtro,
            'fornecedor': fornecedor_filtro,
            'tipo': tipo_filtro,
            'origem': origem_filtro,
            'destino': destino_filtro,
            'data_inicial': data_inicial,
            'data_final': data_final,
        },
    }


def _exportar_ferramenta_relatorio_excel(context):
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo'

    ws_resumo.append(['Indicador', 'Valor'])
    ws_resumo.append(['Modelos', context['total_modelos']])
    ws_resumo.append(['Unidades', context['total_unidades']])
    ws_resumo.append(['No depósito', context['total_deposito']])
    ws_resumo.append(['Em obra', context['total_em_obra']])
    ws_resumo.append(['Em manutenção', context['total_manutencao']])
    ws_resumo.append(['Perdidas', context['total_perdida']])
    ws_resumo.append(['Baixas', context['total_descartada']])

    ws_ferramentas = wb.create_sheet('Ferramentas')
    ws_ferramentas.append(['Código', 'Ferramenta', 'Classificação', 'Fornecedor', 'Total', 'Depósito', 'Obras', 'Manutenção'])
    for f in context['ferramentas_lista']:
        ws_ferramentas.append([
            f.codigo,
            f.nome,
            f.get_classificacao_display(),
            f.fornecedor.nome if f.fornecedor else '-',
            f.quantidade_total,
            f.quantidade_deposito,
            f.quantidade_em_obras,
            f.quantidade_manutencao,
        ])

    ws_obras = wb.create_sheet('Distribuição Obras')
    ws_obras.append(['Obra', 'Quantidade'])
    for item in context['obras_distribuicao_lista']:
        ws_obras.append([item.get('obra__nome') or 'Sem obra', item.get('total') or 0])

    ws_mov = wb.create_sheet('Movimentações')
    ws_mov.append(['Data/Hora', 'Ferramenta', 'Classificação', 'Tipo', 'Quantidade', 'Origem', 'Destino', 'Responsável'])
    for mov in context['movimentacoes_lista']:
        ws_mov.append([
            timezone.localtime(mov.data_movimentacao).strftime('%d/%m/%Y %H:%M'),
            f'{mov.ferramenta.codigo} - {mov.ferramenta.nome}',
            mov.ferramenta.get_classificacao_display(),
            mov.get_tipo_display(),
            mov.quantidade,
            mov.get_origem_label(),
            mov.get_destino_label(),
            mov.responsavel.username if mov.responsavel else '-',
        ])

    for sheet in wb.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio_ferramentas.xlsx"'
    wb.save(response)
    return response


@login_required
def ferramenta_list(request):
    """Lista ferramentas"""
    qs, busca, categoria, classificacao, fornecedor, status = _filtrar_ferramentas_queryset(request)

    # Ordering
    order = request.GET.get('order', 'nome')
    direction = request.GET.get('dir', 'asc')
    allowed = {
        'codigo': 'codigo',
        'nome': 'nome',
        'categoria': 'categoria',
        'quantidade': 'quantidade_total'
    }
    order_field = allowed.get(order, 'nome')
    if direction == 'desc':
        order_field = f'-{order_field}'
    qs = qs.order_by(order_field)

    # Contadores
    base_status_qs = Ferramenta.objects.all()
    localizacao_filters = {}
    if status == 'ativas':
        base_status_qs = base_status_qs.filter(ativo=True)
        localizacao_filters['ferramenta__ativo'] = True
    elif status == 'inativas':
        base_status_qs = base_status_qs.filter(ativo=False)
        localizacao_filters['ferramenta__ativo'] = False

    if categoria:
        base_status_qs = base_status_qs.filter(categoria=categoria)
    if classificacao:
        base_status_qs = base_status_qs.filter(classificacao=classificacao)
    if fornecedor:
        base_status_qs = base_status_qs.filter(fornecedor_id=fornecedor)

    total_ferramentas = base_status_qs.count()
    
    # Somar quantidades por localização
    from django.db.models import Sum
    total_deposito = LocalizacaoFerramenta.objects.filter(
        local_tipo='deposito', **localizacao_filters
    ).aggregate(total=Sum('quantidade'))['total'] or 0
    
    total_em_obra = LocalizacaoFerramenta.objects.filter(
        local_tipo='obra', **localizacao_filters
    ).aggregate(total=Sum('quantidade'))['total'] or 0
    
    total_manutencao = LocalizacaoFerramenta.objects.filter(
        local_tipo='manutencao', **localizacao_filters
    ).aggregate(total=Sum('quantidade'))['total'] or 0
    
    total_resultado = qs.count()

    # Pagination
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    per_page = 10
    paginator = Paginator(qs, per_page)
    page = request.GET.get('page')
    try:
        ferramentas_page = paginator.page(page)
    except PageNotAnInteger:
        ferramentas_page = paginator.page(1)
    except EmptyPage:
        ferramentas_page = paginator.page(paginator.num_pages)

    # Build base querystring (keep filters but remove paging/sorting params)
    params = request.GET.copy()
    for k in ('page', 'order', 'dir', 'per_page'):
        if k in params:
            params.pop(k)
    base_qs = params.urlencode()

    context = {
        'ferramentas': ferramentas_page,
        'title': 'Ferramentas',
        'base_qs': base_qs,
        'current_order': order,
        'current_dir': direction,
        'paginator': paginator,
        'per_page': per_page,
        'category_choices': [(k, v) for k, v in Ferramenta._meta.get_field('categoria').choices],
        'classificacao_choices': Ferramenta.CLASSIFICACAO_CHOICES,
        'fornecedor_choices': Ferramenta.objects.filter(fornecedor__isnull=False).values_list('fornecedor__id', 'fornecedor__nome').distinct().order_by('fornecedor__nome'),
        'busca': busca,
        'categoria_filter': categoria or '',
        'classificacao_filter': classificacao,
        'fornecedor_filter': fornecedor,
        'status_filter': status,
        'total_ferramentas': total_ferramentas,
        'total_deposito': total_deposito,
        'total_em_obra': total_em_obra,
        'total_manutencao': total_manutencao,
        'total_resultado': total_resultado,
    }
    return render(request, 'ferramentas/ferramenta_list.html', context)


@login_required
def ferramenta_relatorio_impressao(request):
    export = request.GET.get('export', '').strip().lower()
    if export in {'pdf', 'xlsx'}:
        context = _build_ferramenta_relatorio_data(request, paginate=False)
        if export == 'xlsx':
            return _exportar_ferramenta_relatorio_excel(context)

        html = render_to_string(
            'ferramentas/ferramenta_relatorio_pdf.html',
            context,
            request=request,
        )
        try:
            from weasyprint import HTML
            pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
        except Exception:
            from io import BytesIO
            from xhtml2pdf import pisa

            pdf_buffer = BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
            if pisa_status.err:
                messages.error(request, 'Não foi possível gerar o PDF deste relatório agora.')
                return render(request, 'ferramentas/ferramenta_relatorio_impressao.html', _build_ferramenta_relatorio_data(request, paginate=True))
            pdf = pdf_buffer.getvalue()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="relatorio_ferramentas.pdf"'
        return response

    context = _build_ferramenta_relatorio_data(request, paginate=True)
    return render(request, 'ferramentas/ferramenta_relatorio_impressao.html', context)

    """
    Relatório imprimível de localização e histórico de ferramentas.
    Se ids forem informados, imprime apenas os selecionados.
    Se não houver ids, usa o filtro atual da listagem.
    """
    ids = request.GET.getlist('ids')
    ferramenta_filtro = request.GET.get('ferramenta', '').strip()
    fornecedor_filtro = request.GET.get('fornecedor', '').strip()
    tipo_filtro = request.GET.get('tipo', '').strip()
    origem_filtro = request.GET.get('origem', '').strip()
    destino_filtro = request.GET.get('destino', '').strip()

    if ids:
        base_ferramentas_qs = Ferramenta.objects.filter(id__in=ids)
    else:
        base_ferramentas_qs, _, _, _, _, _ = _filtrar_ferramentas_queryset(request)

    ferramentas_filtro_opcoes = list(
        base_ferramentas_qs.order_by('nome').values('id', 'codigo', 'nome')
    )

    if ferramenta_filtro:
        base_ferramentas_qs = base_ferramentas_qs.filter(id=ferramenta_filtro)
    if fornecedor_filtro:
        base_ferramentas_qs = base_ferramentas_qs.filter(fornecedor_id=fornecedor_filtro)

    ferramentas_qs = base_ferramentas_qs.order_by('nome')
    ferramentas = list(ferramentas_qs)
    ferramenta_ids = [f.id for f in ferramentas]

    # Totais consolidados
    total_modelos = len(ferramentas)
    total_unidades = sum((f.quantidade_total or 0) for f in ferramentas)
    total_deposito = sum((f.quantidade_deposito or 0) for f in ferramentas)
    total_em_obra = sum((f.quantidade_em_obras or 0) for f in ferramentas)
    total_manutencao = sum((f.quantidade_manutencao or 0) for f in ferramentas)
    total_perdida = sum((f.quantidade_perdida or 0) for f in ferramentas)

    total_descartada = MovimentacaoFerramenta.objects.filter(
        ferramenta_id__in=ferramenta_ids,
        tipo='descarte'
    ).aggregate(total=models.Sum('quantidade'))['total'] or 0

    # Onde estão atualmente (obras)
    obras_distribuicao_qs = (
        LocalizacaoFerramenta.objects
        .filter(ferramenta_id__in=ferramenta_ids, local_tipo='obra', quantidade__gt=0)
        .values('obra_id', 'obra__nome')
        .annotate(total=models.Sum('quantidade'))
        .order_by('obra__nome')
    )

    # Histórico para impressão em lote
    movimentacoes_qs = (
        MovimentacaoFerramenta.objects
        .filter(ferramenta_id__in=ferramenta_ids)
        .select_related('ferramenta', 'ferramenta__fornecedor', 'responsavel', 'obra_origem', 'obra_destino')
        .order_by('-data_movimentacao', '-id')
    )

    if tipo_filtro:
        movimentacoes_qs = movimentacoes_qs.filter(tipo=tipo_filtro)
    if origem_filtro:
        movimentacoes_qs = movimentacoes_qs.filter(
            models.Q(origem_tipo__icontains=origem_filtro) |
            models.Q(obra_origem__nome__icontains=origem_filtro)
        )
    if destino_filtro:
        movimentacoes_qs = movimentacoes_qs.filter(
            models.Q(destino_tipo__icontains=destino_filtro) |
            models.Q(obra_destino__nome__icontains=destino_filtro)
        )

    per_page = 10
    obras_distribuicao = Paginator(obras_distribuicao_qs, per_page).get_page(request.GET.get('page_obra', 1))
    ferramentas = Paginator(ferramentas, per_page).get_page(request.GET.get('page_ferramenta', 1))
    movimentacoes = Paginator(movimentacoes_qs, per_page).get_page(request.GET.get('page_historico', 1))

    def _query_sem(*remove_keys):
        params = request.GET.copy()
        for key in remove_keys:
            params.pop(key, None)
        return params.urlencode()

    context = {
        'title': 'Relatório de Ferramentas',
        'ferramentas': ferramentas,
        'movimentacoes': movimentacoes,
        'obras_distribuicao': obras_distribuicao,
        'tipo_choices': MovimentacaoFerramenta.TIPO_CHOICES,
        'ferramentas_filtro_opcoes': ferramentas_filtro_opcoes,
        'fornecedor_choices': Ferramenta.objects.filter(fornecedor__isnull=False).values_list('fornecedor__id', 'fornecedor__nome').distinct().order_by('fornecedor__nome'),
        'ids_selecionados': ids,
        'qs_sem_page_obras': _query_sem('page_obra'),
        'qs_sem_page_ferramenta': _query_sem('page_ferramenta'),
        'qs_sem_page_historico': _query_sem('page_historico'),
        'total_modelos': total_modelos,
        'total_unidades': total_unidades,
        'total_deposito': total_deposito,
        'total_em_obra': total_em_obra,
        'total_manutencao': total_manutencao,
        'total_perdida': total_perdida,
        'total_descartada': total_descartada,
        'filtros': {
            'q': request.GET.get('q', '').strip(),
            'categoria': request.GET.get('categoria', '').strip(),
            'status': request.GET.get('status', '').strip(),
            'ferramenta': ferramenta_filtro,
            'fornecedor': fornecedor_filtro,
            'tipo': tipo_filtro,
            'origem': origem_filtro,
            'destino': destino_filtro,
        },
    }
    return render(request, 'ferramentas/ferramenta_relatorio_impressao.html', context)


@login_required
def ferramenta_detail(request, pk):
    """Detalhes de uma ferramenta"""
    ferramenta = get_object_or_404(Ferramenta.objects.select_related('fornecedor'), pk=pk)
    movimentacoes_qs = ferramenta.movimentacoes.select_related('responsavel', 'obra_origem', 'obra_destino', 'ferramenta__fornecedor').all()
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(movimentacoes_qs, 10)
    page = request.GET.get('page')
    try:
        movimentacoes = paginator.page(page)
    except PageNotAnInteger:
        movimentacoes = paginator.page(1)
    except EmptyPage:
        movimentacoes = paginator.page(paginator.num_pages)
    context = {
        'ferramenta': ferramenta,
        'movimentacoes': movimentacoes,
        'title': ferramenta.nome
    }
    return render(request, 'ferramentas/ferramenta_detail.html', context)


@login_required
def ferramenta_create(request):
    """Cadastra nova ferramenta"""
    if request.method == 'POST':
        form = FerramentaForm(request.POST, request.FILES)
        if form.is_valid():
            # If codigo missing, generate a unique code
            codigo = form.cleaned_data.get('codigo')
            if not codigo:
                prefixes = ['FRR', 'TLS', 'SEG', 'MED', 'OTR']
                tentativa = 0
                codigo = None
                existing = set(Ferramenta.objects.values_list('codigo', flat=True))
                while tentativa < 20:
                    prefix = random.choice(prefixes)
                    numero = random.randint(10000, 99999)
                    candidate = f"{prefix}-{numero}"
                    if candidate not in existing:
                        codigo = candidate
                        break
                    tentativa += 1
                if not codigo:
                    # fallback to timestamp-based code
                    codigo = f"FERR-{int(timezone.now().timestamp())}"
                # inject into form instance
                form.instance.codigo = codigo

            ferramenta = form.save(commit=False)
            
            # Guardar quantidade informada
            quantidade_inicial = ferramenta.quantidade_total
            
            # Zerar quantidade_total (será incrementada pela movimentação)
            ferramenta.quantidade_total = 0
            ferramenta.save()
            
            # Registrar entrada no depósito (isso cria LocalizacaoFerramenta e incrementa quantidade_total)
            if quantidade_inicial > 0:
                MovimentacaoFerramenta.objects.create(
                    ferramenta=ferramenta,
                    quantidade=quantidade_inicial,
                    tipo='entrada_deposito',
                    origem_tipo='compra',
                    destino_tipo='deposito',
                    responsavel=request.user,
                    observacoes='Entrada inicial criada automaticamente no cadastro da ferramenta.'
                )
            
            messages.success(request, f'Ferramenta {ferramenta.codigo} criada com sucesso.')
            return redirect('ferramentas:ferramenta_detail', pk=ferramenta.pk)
    else:
        form = FerramentaForm()

    return render(request, 'ferramentas/ferramenta_form.html', {'form': form, 'title': 'Nova Ferramenta', 'classificacao_choices': Ferramenta.CLASSIFICACAO_CHOICES})


@login_required
def ferramenta_update(request, pk):
    """Edita uma ferramenta existente"""
    ferramenta = get_object_or_404(Ferramenta, pk=pk)
    if request.method == 'POST':
        form = FerramentaForm(request.POST, request.FILES, instance=ferramenta)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ferramenta atualizada.')
            return redirect('ferramentas:ferramenta_detail', pk=ferramenta.pk)
    else:
        form = FerramentaForm(instance=ferramenta)

    return render(request, 'ferramentas/ferramenta_form.html', {'form': form, 'ferramenta': ferramenta, 'title': 'Editar Ferramenta', 'classificacao_choices': Ferramenta.CLASSIFICACAO_CHOICES})


@login_required
def movimentacao_create(request):
    """Registra movimentação de ferramenta"""
    if request.method == 'POST':
        itens_json = request.POST.get('itens_json', '').strip()
        if itens_json:
            try:
                itens = json.loads(itens_json)
            except json.JSONDecodeError:
                messages.error(request, 'Lista de movimentações inválida.')
                return render(request, 'ferramentas/movimentacao_form.html', {'form': MovimentacaoForm(), 'title': 'Movimentar Ferramenta'})

            if not isinstance(itens, list) or not itens:
                messages.error(request, 'Adicione ao menos uma movimentação antes de salvar.')
                return render(request, 'ferramentas/movimentacao_form.html', {'form': MovimentacaoForm(), 'title': 'Movimentar Ferramenta'})

            with transaction.atomic():
                total_criadas = 0
                for idx, item in enumerate(itens, start=1):
                    form_item = MovimentacaoForm(item)
                    if not form_item.is_valid():
                        erro = '; '.join(
                            f'{campo}: {", ".join(msgs)}'
                            for campo, msgs in form_item.errors.items()
                        )
                        messages.error(request, f'Erro no item {idx}: {erro}')
                        return render(
                            request,
                            'ferramentas/movimentacao_form.html',
                            {'form': MovimentacaoForm(), 'title': 'Movimentar Ferramenta'}
                        )

                    mov = form_item.save(commit=False)
                    mov.responsavel = request.user
                    mov.save()
                    total_criadas += 1

            messages.success(request, f'{total_criadas} movimentação(ões) registrada(s).')
            return redirect('ferramentas:ferramenta_list')

        form = MovimentacaoForm(request.POST)
        if form.is_valid():
            mov = form.save(commit=False)
            mov.responsavel = request.user
            mov.save()
            messages.success(request, 'Movimentação registrada.')
            return redirect('ferramentas:ferramenta_detail', pk=mov.ferramenta.pk)
    else:
        # allow preselecting ferramenta via GET param ?f=<pk>
        f_pk = request.GET.get('f')
        if f_pk:
            form = MovimentacaoForm(initial={'ferramenta': f_pk})
        else:
            form = MovimentacaoForm()

    return render(request, 'ferramentas/movimentacao_form.html', {'form': form, 'title': 'Movimentar Ferramenta'})


@login_required
def conferencia_list(request):
    """Lista conferências de ferramentas"""
    conferencias = ConferenciaFerramenta.objects.all().select_related('obra', 'fiscal')
    context = {
        'conferencias': conferencias,
        'title': 'Conferências'
    }
    return render(request, 'ferramentas/conferencia_list.html', context)


@login_required
def conferencia_detail(request, pk):
    conf = get_object_or_404(ConferenciaFerramenta, pk=pk)
    itens = conf.itens.select_related('ferramenta').all()
    context = {
        'conferencia': conf,
        'itens': itens,
        'title': f'Conferência - {conf.obra.nome}'
    }
    return render(request, 'ferramentas/conferencia_detail.html', context)


@login_required
def conferencia_create(request):
    """
    Cria conferência e AUTOMATICAMENTE popula com as ferramentas da obra.
    Não precisa mais adicionar itens manualmente!
    """
    if request.method == 'POST':
        form = ConferenciaForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                # Salvar conferência
                conf = form.save(commit=False)
                conf.fiscal = request.user
                conf.data_conferencia = timezone.localtime()
                conf.save()
                
                # BUSCAR AUTOMATICAMENTE ferramentas que DEVERIAM estar na obra
                obra = conf.obra
                
                # Ferramentas que estão nesta obra (usando LocalizacaoFerramenta)
                localizacoes_na_obra = LocalizacaoFerramenta.objects.filter(
                    local_tipo='obra',
                    obra=obra,
                    quantidade__gt=0  # Apenas localizações com quantidade > 0
                ).select_related('ferramenta')
                
                # CRIAR AUTOMATICAMENTE os itens de conferência
                itens_criados = 0
                for loc in localizacoes_na_obra:
                    ItemConferencia.objects.create(
                        conferencia=conf,
                        ferramenta=loc.ferramenta,
                        quantidade_esperada=loc.quantidade,  # Quantidade que deveria ter
                        quantidade_encontrada=0,  # Fiscal vai preencher
                        status='',  # Será calculado automaticamente no save()
                        observacoes=''
                    )
                    itens_criados += 1
                
                if itens_criados == 0:
                    messages.warning(
                        request,
                        f'Conferência criada, mas não há ferramentas registradas em {obra.nome}. '
                        f'Verifique se as ferramentas foram movimentadas corretamente.'
                    )
                else:
                    messages.success(
                        request, 
                        f'✅ Conferência criada com {itens_criados} ferramenta(s) para conferir!'
                    )
                
                # Redirecionar para página de CONFERIR (não de listar)
                return redirect('ferramentas:conferencia_conferir', pk=conf.pk)
    else:
        form = ConferenciaForm()
    
    # Adicionar informações úteis ao contexto
    obras_com_ferramentas = {}
    for obra in Obra.objects.filter(ativo=True):
        qtd = LocalizacaoFerramenta.objects.filter(
            local_tipo='obra',
            obra=obra,
            quantidade__gt=0
        ).count()
        if qtd > 0:
            obras_com_ferramentas[obra.id] = qtd
    
    return render(request, 'ferramentas/conferencia_form.html', {
        'form': form,
        'title': 'Nova Conferência',
        'obras_com_ferramentas': obras_com_ferramentas
    })


@login_required
def conferencia_conferir(request, pk):
    """
    Tela onde o fiscal confere as ferramentas.
    Mostra lista de TODAS as ferramentas que deveriam estar na obra.
    Fiscal marca quantidade encontrada e o sistema calcula o status automaticamente.
    """
    conferencia = get_object_or_404(
        ConferenciaFerramenta.objects.select_related('obra', 'fiscal'),
        pk=pk
    )
    itens = conferencia.itens.select_related('ferramenta').order_by('ferramenta__nome')
    
    if request.method == 'POST':
        # Processar formulário de conferência
        with transaction.atomic():
            itens_processados = 0
            
            for item in itens:
                qtd_key = f'quantidade_encontrada_{item.id}'
                obs_key = f'obs_{item.id}'
                
                # Pegar quantidade encontrada
                try:
                    qtd_encontrada = int(request.POST.get(qtd_key, 0))
                    if qtd_encontrada < 0:
                        qtd_encontrada = 0
                except (ValueError, TypeError):
                    qtd_encontrada = 0
                
                item.quantidade_encontrada = qtd_encontrada
                item.observacoes = request.POST.get(obs_key, '').strip()
                
                # O status é calculado automaticamente no save() do model
                item.save()
                itens_processados += 1
                
                # Se FALTOU ferramentas, criar observação automática
                if item.status == 'falta' and not item.observacoes:
                    diferenca = abs(item.diferenca)
                    item.observacoes = f'Faltam {diferenca} unidade(s)'
                    item.save(update_fields=['observacoes'])
                
                # Se SOBROU ferramentas, criar observação automática
                elif item.status == 'sobra' and not item.observacoes:
                    diferenca = item.diferenca
                    item.observacoes = f'Sobraram {diferenca} unidade(s) não registradas'
                    item.save(update_fields=['observacoes'])
            
            # Atualizar observações gerais da conferência
            obs_gerais = request.POST.get('observacoes_gerais', '').strip()
            if obs_gerais:
                conferencia.observacoes_gerais = obs_gerais
                conferencia.save(update_fields=['observacoes_gerais'])
            
            messages.success(
                request, 
                f'✅ Conferência salva com sucesso! {itens_processados} item(ns) conferido(s).'
            )
            
            # Mostrar alerta se há divergências
            if conferencia.tem_divergencias:
                messages.warning(
                    request,
                    f'⚠️ Atenção: {conferencia.total_divergencias} divergência(s) encontrada(s)!'
                )
            
            return redirect('ferramentas:conferencia_detail', pk=conferencia.pk)
    
    # Estatísticas para a tela
    total_esperado = sum(item.quantidade_esperada for item in itens)
    
    context = {
        'conferencia': conferencia,
        'itens': itens,
        'total_esperado': total_esperado,
        'title': f'Conferir Ferramentas - {conferencia.obra.nome}'
    }
    return render(request, 'ferramentas/conferencia_conferir.html', context)


@login_required
def conferencia_list(request):
    """Lista conferências de ferramentas com informações de divergências"""
    conferencias = ConferenciaFerramenta.objects.all().select_related(
        'obra', 'obra__cliente', 'fiscal'
    ).prefetch_related('itens')
    
    # Filtros
    obra = request.GET.get('obra', '').strip()
    cliente = request.GET.get('cliente', '').strip()
    fiscal = request.GET.get('fiscal', '').strip()
    data = request.GET.get('data', '').strip()
    status = request.GET.get('status', '').strip()

    if obra:
        conferencias = conferencias.filter(obra__nome__icontains=obra)

    if cliente:
        conferencias = conferencias.filter(obra__cliente__nome__icontains=cliente)

    if fiscal:
        conferencias = conferencias.filter(
            models.Q(fiscal__first_name__icontains=fiscal) |
            models.Q(fiscal__last_name__icontains=fiscal) |
            models.Q(fiscal__username__icontains=fiscal)
        )

    if data:
        conferencias = conferencias.filter(data_conferencia__date=data)
    
    # Adicionar informações de divergência
    conferencias_com_info = []
    for conf in conferencias:
        info = {
            'conferencia': conf,
            'total_itens': conf.total_itens,
            'total_divergencias': conf.total_divergencias,
            'tem_divergencias': conf.tem_divergencias
        }
        if status == 'ok' and info['tem_divergencias']:
            continue
        if status == 'divergencia' and not info['tem_divergencias']:
            continue
        conferencias_com_info.append(info)

    paginator = Paginator(conferencias_com_info, 10)
    conferencias_page = paginator.get_page(request.GET.get('page'))

    params = request.GET.copy()
    if 'page' in params:
        params.pop('page')
    base_qs = params.urlencode()

    context = {
        'conferencias_info': conferencias_page,
        'base_qs': base_qs,
        'filtro_obra': obra,
        'filtro_cliente': cliente,
        'filtro_fiscal': fiscal,
        'filtro_data': data,
        'filtro_status': status,
        'title': 'Conferências de Ferramentas'
    }
    return render(request, 'ferramentas/conferencia_list.html', context)


class ConferenciaCreateView(LoginRequiredMixin, generic.CreateView):
    model = ConferenciaFerramenta
    form_class = ConferenciaForm
    template_name = 'ferramentas/conferencia_form.html'
    success_url = reverse_lazy('ferramentas:conferencia_list')

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.fiscal = self.request.user
        obj.data_conferencia = timezone.localtime()
        obj.save()
        messages.success(self.request, 'Conferência criada.')
        return super().form_valid(form)


class ItemConferenciaCreateView(LoginRequiredMixin, generic.CreateView):
    model = ItemConferencia
    form_class = ItemConferenciaForm
    template_name = 'ferramentas/itemconferencia_form.html'

    def get_initial(self):
        initial = super().get_initial()
        conferencia_pk = self.kwargs.get('conferencia_pk')
        if conferencia_pk:
            initial['conferencia'] = conferencia_pk
        return initial
    
    def get_form_kwargs(self):
        """Passa a obra para o form se houver conferência"""
        kwargs = super().get_form_kwargs()
        conferencia_pk = self.kwargs.get('conferencia_pk')
        if conferencia_pk:
            try:
                conf = ConferenciaFerramenta.objects.get(pk=conferencia_pk)
                kwargs['obra'] = conf.obra
            except ConferenciaFerramenta.DoesNotExist:
                pass
        return kwargs

    def form_valid(self, form):
        obj = form.save(commit=False)
        if not obj.conferencia_id:
            obj.conferencia = get_object_or_404(ConferenciaFerramenta, pk=self.kwargs.get('conferencia_pk'))
        obj.save()
        messages.success(self.request, 'Item adicionado à conferência.')
        return redirect('ferramentas:conferencia_list')


class MovimentacaoCreateView(LoginRequiredMixin, generic.CreateView):
    model = MovimentacaoFerramenta
    form_class = MovimentacaoForm
    template_name = 'ferramentas/movimentacao_form.html'

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.responsavel = self.request.user
        obj.save()
        messages.success(self.request, 'Movimentação registrada.')
        return redirect('ferramentas:ferramenta_detail', pk=obj.ferramenta.pk)

    def get_initial(self):
        initial = super().get_initial()
        f_pk = self.request.GET.get('f')
        if f_pk:
            initial['ferramenta'] = f_pk
        return initial


class ConferenciaWithItemsCreateView(LoginRequiredMixin, View):
    """Create a Conferencia and multiple ItemConferencia at once using an inline formset."""
    template_name = 'ferramentas/conferencia_form_with_items.html'
    form_class = ConferenciaForm
    ItemFormSet = inlineformset_factory(
        parent_model=ConferenciaFerramenta,
        model=ItemConferencia,
        form=ItemConferenciaForm,
        extra=8,
        can_delete=True
    )

    def get(self, request, *args, **kwargs):
        form = self.form_class()
        formset = self.ItemFormSet()
        return render(request, self.template_name, {'form': form, 'formset': formset, 'title': 'Nova Conferência (várias ferramentas)'} )

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        formset = self.ItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                conferencia = form.save(commit=False)
                conferencia.fiscal = request.user
                conferencia.data_conferencia = timezone.localtime()
                conferencia.save()
                
                # Salvar itens com quantidade_esperada preenchida
                items = formset.save(commit=False)
                for item in items:
                    item.conferencia = conferencia
                    # Auto-preencher quantidade_esperada se não preenchido
                    if not item.quantidade_esperada and item.ferramenta:
                        try:
                            loc = item.ferramenta.localizacoes.get(
                                local_tipo='obra',
                                obra=conferencia.obra
                            )
                            item.quantidade_esperada = loc.quantidade
                        except LocalizacaoFerramenta.DoesNotExist:
                            item.quantidade_esperada = 0
                    item.save()
                
                for obj in formset.deleted_objects:
                    obj.delete()
            messages.success(request, 'Conferência e itens salvos com sucesso.')
            return redirect('ferramentas:conferencia_list')

        return render(request, self.template_name, {'form': form, 'formset': formset, 'title': 'Nova Conferência (várias ferramentas)'} )


class ConferenciaItemsManageView(LoginRequiredMixin, View):
    """Manage items for an existing Conferencia: show existing items and allow adding/removing many at once."""
    template_name = 'ferramentas/conferencia_form_with_items.html'
    form_class = ConferenciaForm
    ItemFormSet = inlineformset_factory(
        parent_model=ConferenciaFerramenta,
        model=ItemConferencia,
        form=ItemConferenciaForm,
        extra=3,
        can_delete=True
    )

    def get_object(self, pk):
        try:
            return ConferenciaFerramenta.objects.get(pk=pk)
        except ConferenciaFerramenta.DoesNotExist:
            raise Http404()

    def get(self, request, conferencia_pk, *args, **kwargs):
        conferencia = self.get_object(conferencia_pk)
        form = self.form_class(instance=conferencia)
        formset = self.ItemFormSet(instance=conferencia)
        return render(request, self.template_name, {'form': form, 'formset': formset, 'title': f'Conferência {conferencia.obra.nome} - Itens'})

    def post(self, request, conferencia_pk, *args, **kwargs):
        conferencia = self.get_object(conferencia_pk)
        form = self.form_class(request.POST, instance=conferencia)
        formset = self.ItemFormSet(request.POST, instance=conferencia)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                instances = formset.save(commit=False)
                for inst in instances:
                    inst.conferencia = conferencia
                    inst.save()
                # handle deletions
                for obj in formset.deleted_objects:
                    obj.delete()
            messages.success(request, 'Conferência e itens atualizados com sucesso.')
            return redirect('ferramentas:conferencia_detail', pk=conferencia.pk)

        return render(request, self.template_name, {'form': form, 'formset': formset, 'title': f'Conferência {conferencia.obra.nome} - Itens'})

    # end of ConferenciaItemsManageView
