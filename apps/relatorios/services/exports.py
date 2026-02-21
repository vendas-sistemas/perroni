"""
Camada de exportação – PDF (reportlab.platypus) e Excel (openpyxl).
"""

import io
import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)


def _format_brl(value):
    """Format number to Brazilian currency style: R$ 1.234,56"""
    try:
        v = float(value)
    except Exception:
        return str(value)
    s = f"{v:,.2f}"  # 1,234.56
    # swap separators to Brazilian format 1.234,56
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"R$ {s}"

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.relatorios.services.analytics_indicadores import gerar_relatorio_completo_indicadores


# ═══════════════════════════════════════════
#  PDF
# ═══════════════════════════════════════════

def _header_style():
    styles = getSampleStyleSheet()
    return ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        spaceAfter=10,
        spaceBefore=18,
        textColor=colors.HexColor('#1a237e'),
    )


def _table_style():
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ])


def exportar_pdf(filtros: dict | None = None) -> io.BytesIO:
    """Gera relatório completo em PDF e retorna BytesIO."""
    dados = gerar_relatorio_completo_indicadores(filtros)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    h_style = _header_style()
    t_style = _table_style()
    elements = []

    # ── Título ──
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=6,
    )
    elements.append(Paragraph('Relatório de Produção Diária', title_style))
    elements.append(Paragraph(
        f'Gerado em {datetime.datetime.now():%d/%m/%Y %H:%M}',
        styles['Normal'],
    ))
    elements.append(Spacer(1, 12))

    # ── Seção 1: Ranking por Etapa ──
    elements.append(Paragraph('1. Ranking por Etapa (Melhores e Piores)', h_style))

    # O serviço `gerar_relatorio_completo_indicadores` retorna a chave
    # `ranking_por_etapas` contendo uma lista de etapas, cada uma com
    # uma lista de `indicadores` que por sua vez contém `melhores` e
    # `piores`. Adaptamos a exportação para essa estrutura.
    for etapa in dados.get('ranking_por_etapas', []):
        etapa_nome = etapa.get('nome') or etapa.get('etapa_nome') or 'Etapa'
        elements.append(Paragraph(etapa_nome, styles['Heading3']))

        for indicador in etapa.get('indicadores', []):
            # Cabeçalho do indicador
            elements.append(Paragraph(f"📊 {indicador.get('nome', '')} ({indicador.get('unidade','')})", styles['Heading4']))

            # Melhores
            elements.append(Paragraph('Top 3 Melhores', styles['Heading5']))
            header = ['#', 'Pedreiro', f"Média {indicador.get('unidade','')}/dia", 'Dias Trab.']
            rows = [header]
            for i, m in enumerate(indicador.get('melhores', []), 1):
                rows.append([str(i), m.get('nome', ''), f"{float(m.get('media_producao', 0)):.2f}", str(m.get('total_dias', ''))])
            t = Table(rows, repeatRows=1)
            t.setStyle(t_style)
            elements.append(t)
            elements.append(Spacer(1, 6))

            # Piores
            elements.append(Paragraph('Top 3 Piores', styles['Heading5']))
            rows_p = [header]
            for i, p in enumerate(indicador.get('piores', []), 1):
                rows_p.append([str(i), p.get('nome', ''), f"{float(p.get('media_producao', 0)):.2f}", str(p.get('total_dias', ''))])
            t2 = Table(rows_p, repeatRows=1)
            t2.setStyle(t_style)
            elements.append(t2)
            elements.append(Spacer(1, 10))

    elements.append(PageBreak())

    # ── Seção 2: Média de Dias por Etapa ──
    elements.append(Paragraph('2. Média de Dias para Execução de Cada Etapa', h_style))
    header2 = ['Etapa', 'Média de Dias', 'Obras Analisadas']
    rows2 = [header2]
    for e in dados['media_dias_etapa']:
        rows2.append([e['etapa_nome'], f"{e['media_dias']:.1f}", str(e['total_obras'])])
    t3 = Table(rows2, repeatRows=1)
    t3.setStyle(t_style)
    elements.append(t3)
    elements.append(Spacer(1, 20))

    # ── Seção 3: Média Individual ──
    elements.append(Paragraph('3. Média de Rendimento Individual', h_style))
    header3 = ['#', 'Pedreiro', 'Média de Produção/dia (todas etapas)', 'Dias Trabalhados', 'Dias c/ Ociosidade', 'Dias c/ Retrabalho', 'Horas Trabalhadas', 'Total Recebido (soma diárias)', 'Metragem Executada (m²)']
    rows3 = [header3]
    for i, r in enumerate(dados['media_individual'], 1):
        rows3.append([
            str(i),
            r['nome'],
            f"{r['media_producao']:.2f}",
            str(r['total_dias']),
            str(r['total_ociosidade']),
            str(r['total_retrabalho']),
            f"{float(r.get('total_horas', 0)):.1f}",
            _format_brl(r.get('total_valor', 0)),
            f"{float(r.get('total_metragem', 0)):.2f}",
        ])
    t4 = Table(rows3, repeatRows=1)
    t4.setStyle(t_style)
    # Alinhar coluna Valor Total (índice 7) à direita
    t4.setStyle(TableStyle([('ALIGN', (7, 1), (7, -1), 'RIGHT')]))
    elements.append(t4)

    doc.build(elements)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════
#  EXCEL
# ═══════════════════════════════════════════

_HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
_HEADER_FILL = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_CELL_ALIGN = Alignment(horizontal='center', vertical='center')
_THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ''
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def exportar_excel(filtros: dict | None = None) -> io.BytesIO:
    """Gera relatório completo em Excel e retorna BytesIO."""
    dados = gerar_relatorio_completo_indicadores(filtros)
    wb = openpyxl.Workbook()

    # ── Aba 1: Ranking ──
    ws1 = wb.active
    ws1.title = 'Ranking por Etapa'
    row = 1

    # Percorre etapas e indicadores (estrutura retornada por gerar_relatorio_completo_indicadores)
    for etapa in dados.get('ranking_por_etapas', []):
        etapa_nome = etapa.get('nome') or etapa.get('etapa_nome') or 'Etapa'
        # Título da etapa
        ws1.cell(row=row, column=1, value=etapa_nome).font = Font(bold=True, size=12)
        row += 1

        for indicador in etapa.get('indicadores', []):
            indicador_nome = indicador.get('nome', '')
            indicador_unidade = indicador.get('unidade', '')

            # Cabeçalho do indicador
            ws1.cell(row=row, column=1, value=f"{indicador_nome} ({indicador_unidade})").font = Font(bold=True)
            row += 1

            # Melhores
            ws1.cell(row=row, column=1, value='Top 3 Melhores').font = Font(bold=True, color='2E7D32')
            row += 1
            headers = ['#', 'Pedreiro', f'Média {indicador_unidade}/dia', 'Dias Trab.']
            _write_header(ws1, headers, row)
            row += 1
            for i, m in enumerate(indicador.get('melhores', []), 1):
                ws1.cell(row=row, column=1, value=i).alignment = _CELL_ALIGN
                ws1.cell(row=row, column=2, value=m.get('nome')).alignment = _CELL_ALIGN
                ws1.cell(row=row, column=3, value=m.get('media_producao')).alignment = _CELL_ALIGN
                ws1.cell(row=row, column=4, value=m.get('total_dias')).alignment = _CELL_ALIGN
                for c in range(1, 5):
                    ws1.cell(row=row, column=c).border = _THIN_BORDER
                row += 1

            # Piores
            ws1.cell(row=row, column=1, value='Top 3 Piores').font = Font(bold=True, color='C62828')
            row += 1
            _write_header(ws1, headers, row)
            row += 1
            for i, p in enumerate(indicador.get('piores', []), 1):
                ws1.cell(row=row, column=1, value=i).alignment = _CELL_ALIGN
                ws1.cell(row=row, column=2, value=p.get('nome')).alignment = _CELL_ALIGN
                ws1.cell(row=row, column=3, value=p.get('media_producao')).alignment = _CELL_ALIGN
                ws1.cell(row=row, column=4, value=p.get('total_dias')).alignment = _CELL_ALIGN
                for c in range(1, 5):
                    ws1.cell(row=row, column=c).border = _THIN_BORDER
                row += 1

            row += 1  # linha em branco

    _auto_width(ws1)

    # ── Aba 2: Média por Etapa ──
    ws2 = wb.create_sheet('Média por Etapa')
    headers2 = ['Etapa', 'Média de Dias', 'Obras Analisadas']
    _write_header(ws2, headers2)
    for i, e in enumerate(dados['media_dias_etapa'], 2):
        ws2.cell(row=i, column=1, value=e['etapa_nome']).alignment = _CELL_ALIGN
        ws2.cell(row=i, column=2, value=e['media_dias']).alignment = _CELL_ALIGN
        ws2.cell(row=i, column=3, value=e['total_obras']).alignment = _CELL_ALIGN
        for c in range(1, 4):
            ws2.cell(row=i, column=c).border = _THIN_BORDER
    _auto_width(ws2)

    # ── Aba 3: Média Individual ──
    ws3 = wb.create_sheet('Média Individual')
    headers3 = ['#', 'Pedreiro', 'Média de Produção/dia (todas etapas)', 'Dias Trabalhados', 'Dias c/ Ociosidade', 'Dias c/ Retrabalho', 'Horas Trabalhadas', 'Total Recebido (soma diárias)', 'Metragem Executada (m²)']
    _write_header(ws3, headers3)
    for i, r in enumerate(dados['media_individual'], 2):
        ws3.cell(row=i, column=1, value=i - 1).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=2, value=r['nome']).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=3, value=r.get('media_producao')).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=4, value=r.get('total_dias')).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=5, value=r.get('total_ociosidade')).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=6, value=r.get('total_retrabalho')).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=7, value=r.get('total_horas')).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=8, value=r.get('total_valor')).alignment = _CELL_ALIGN
        ws3.cell(row=i, column=9, value=r.get('total_metragem')).alignment = _CELL_ALIGN
        for c in range(1, 10):
            ws3.cell(row=i, column=c).border = _THIN_BORDER
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
